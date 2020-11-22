# -*- coding: utf-8 -*-
import fileinput
import os
import pickle
import re
import logging
from scipy.stats import spearmanr
from gensim.models import KeyedVectors, Word2Vec
from openpyxl import load_workbook
from itertools import product

globalnonExistingId = 999999999

logging.basicConfig(level=logging.ERROR,
                    format='%(asctime)s [line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt='%a, %d %b %Y %H:%M:%S')
logger = logging.getLogger()
logger.setLevel(logging.ERROR)

class DataSet:

    def __init__(self, excel_score_path, concept_path):
        assert os.path.isfile(excel_score_path), u'{} is not a real existing file!'.format(excel_score_path)
        assert os.path.isfile(concept_path), u'{} is not a real existing file!'.format(concept_path)
        self.excel_score_path = excel_score_path
        self.concept_path = concept_path
        self.wb = load_workbook(self.excel_score_path)

        self.wordpairs = []  # 数据集中所有要计算的词语对
        self.dict_wordpair2humanscore = {}
        self.word_concept = {}  # 保存了word对应的完整概念信息
        self.valid_word_concept = {}  # 仅保留当word对应的概念在词向量中存在的概念
        self.title2vecindex = {}  # 保存title到词向量中索引的映射
        self.result = {}

        self.Title_Redirect = [u'Title', u'RedirectTitle']
        self.DisambigItems = [u'Title', u'RedirectTitle', u'DisambigItems']
        self.HatnoteLinkItems = [u'Title', u'RedirectTitle', u'HatnoteLinkItems', u'DisambigItems']
        self.HatnoteLinkDisambig = [u'Title', u'RedirectTitle', u'HatnoteLinkItems', u'DisambigItems',
                               u'HatnoteLinkDisambig']
        self.HatnoteLinkDisambigItems = [u'Title', u'RedirectTitle', u'HatnoteLinkItems', u'DisambigItems',
                                    u'HatnoteLinkDisambig', u'HatnoteLinkDisambigItems']

        self.load_excel()
        self.load_concept()

    def load_excel(self):
        ws = self.wb.active

        # newcolumnid = ws.max_column + 1
        dict_string2rowid = {}

        # dict_wordpair2autorunscore = {}
        for rowid in range(1, ws.max_row + 1):
            a = ws.cell(row=rowid, column=1).value
            b = ws.cell(row=rowid, column=2).value
            c = ws.cell(row=rowid, column=3).value
            if a == u'Word 1' and b == u'Word 2' and c == u'Human (mean)':
                continue
            if a is None and b is None and c is not None:
                dict_string2rowid[c] = rowid
            elif a is not None and b is not None and c is not None:
                key = (a,b)
                dict_string2rowid[key] = rowid
                self.dict_wordpair2humanscore[key] = c
                self.wordpairs.append(key)
            else:
                assert False, '{}: need to update the source code to handle this'

    def load_concept(self):
        concept_file = fileinput.FileInput(self.concept_path)
        re_tag = re.compile(
            r"^Title:(.*?) {4}bNormal:(.*?) {4}bDisambig:(.*?) {4}bRedirect:(.*?) {4}RedirectTitle:(.*?) {4}bPageErr:(.*?) {4}bHttpTimeOut:(.*?) {4}DisambigItems:(.*?) {4}HatnoteLinkItems:(.*?) {4}HatnoteLinkDisambig:(.*?) {4}HatnoteLinkDisambigItems:(.*?)$")
        for line_num, line in enumerate(concept_file):
            line = line.strip().decode('utf-8')
            m = re_tag.match(line)
            assert m is not None, u'the {} line of "{}" is wrong!: {}'.format(line_num+1, self.concept_path, line)
            # 从文件中读出，存入变量
            Title = m.group(1)

            if m.group(2) == u'True':
                bNormal = True
            elif m.group(2) == u'False':
                bNormal = False
            else:
                assert False, u'bNormal: {} need to update handle code'.format(m.group(2))

            if m.group(3) == u'True':
                bDisambig = True
            elif m.group(3) == u'None':
                bDisambig = None
            else:
                assert False, u'bDisambig: {} need to update handle code'.format(m.group(3))

            if m.group(4) == u'None':
                bRedirect = None
            elif m.group(4) == u'True':
                bRedirect = True
            else:
                assert False, u'bRedirect: {} need to update handle code'.format(m.group(4))

            if m.group(5) == u'None':
                RedirectTitle = None
            else:
                RedirectTitle = m.group(5)

            if m.group(6) == u'None':
                bPageErr = None
            elif m.group(6) == u'True':
                bPageErr = True
            else:
                assert False, u'bPageErr: {} need to update handle code'.format(m.group(6))

            if m.group(7) == u'None':
                bHttpTimeOut = None
            elif m.group(7) == u'True':
                bHttpTimeOut = True
            else:
                assert False, u'bHttpTimeOut: {} need to update handle code'.format(m.group(7))

            if m.group(8) == u'None':
                DisambigItems = None
            else:
                DisambigItems = m.group(8).split(u'\t')

            if m.group(9) == u'None':
                HatnoteLinkItems = None
            else:
                HatnoteLinkItems = m.group(9).split(u'\t')

            if m.group(10) == u'None':
                HatnoteLinkDisambig = None
            else:
                HatnoteLinkDisambig = m.group(10).split(u'\t')

            if m.group(11) == u'None':
                HatnoteLinkDisambigItems = None
            else:
                HatnoteLinkDisambigItems = m.group(11).split(u'\t')

            # 将信息存入词典，以备以后使用
            if Title not in self.word_concept:
                conceptinfo = {u'Title': Title, u'bNormal': bNormal, u'bDisambig': bDisambig, u'bRedirect': bRedirect,
                               u'RedirectTitle': RedirectTitle, u'bPageErr': bPageErr, u'bHttpTimeOut': bHttpTimeOut,
                               u'DisambigItems': DisambigItems, u'HatnoteLinkItems': HatnoteLinkItems,
                               u'HatnoteLinkDisambig': HatnoteLinkDisambig,
                               u'HatnoteLinkDisambigItems': HatnoteLinkDisambigItems}

                self.word_concept[Title] = conceptinfo

        concept_file.close()

    def valid_concept(self, model):
        """
        验证各个单词的候选概念集合，只保留在词向量中出现的部分
        :return:
        """
        self.valid_word_concept = {}
        for word, concepts in self.word_concept.iteritems():
            self.valid_word_concept[word] = {}
            for k,v in concepts.iteritems():
                if isinstance(v, list):
                    new = []
                    for title in v:
                        if title in self.title2vecindex:
                            new.append(self.title2vecindex[title])
                        else:
                            vec_index = model.get_vec_index(title)
                            if vec_index is not None:
                                self.title2vecindex[title] = vec_index
                                new.append(vec_index)
                    self.valid_word_concept[word][k] = new
                elif isinstance(v, unicode):
                    if v in self.title2vecindex:
                        self.valid_word_concept[word][k] = self.title2vecindex[v]
                    else:
                        vec_index = model.get_vec_index(v)
                        if vec_index is not None:
                            self.title2vecindex[v] = vec_index
                            self.valid_word_concept[word][k] = vec_index
                        else:
                            self.valid_word_concept[word][k] = u''
                else:
                    self.valid_word_concept[word][k] = v

    def save_valid_concept(self, file_path):
        line = u'Key:{}    Title:{}    RedirectTitle:{}    bPageErr:{}    DisambigItems:{}    HatnoteLinkItems:{}    HatnoteLinkDisambig:{}    HatnoteLinkDisambigItems:{}\n'
        with open(file_path, 'w') as f:
            for word, concepts in self.valid_word_concept.iteritems():
                Title = self.reduce_output(concepts[u'Title'])
                RedirectTitle = self.reduce_output(concepts[u'RedirectTitle'])
                bPageErr = self.reduce_output(concepts[u'bPageErr'])
                DisambigItems = self.reduce_output(concepts[u'DisambigItems'], u'\t')
                HatnoteLinkItems = self.reduce_output(concepts[u'HatnoteLinkItems'], u'\t')
                HatnoteLinkDisambig = self.reduce_output(concepts[u'HatnoteLinkDisambig'], u'\t')
                HatnoteLinkDisambigItems = self.reduce_output(concepts[u'HatnoteLinkDisambigItems'], u'\t')
                f.write(line.format(word, Title, RedirectTitle, bPageErr, DisambigItems, HatnoteLinkItems, HatnoteLinkDisambig, HatnoteLinkDisambigItems).encode('utf-8'))

    def load_valid_concept(self, file_path):
        concept_file = fileinput.FileInput(file_path)
        re_tag = re.compile(r'^Key:(.*?) {4}Title:(.*?) {4}RedirectTitle:(.*?) {4}bPageErr:(.*?) {4}DisambigItems:(.*?) {4}HatnoteLinkItems:(.*?) {4}HatnoteLinkDisambig:(.*?) {4}HatnoteLinkDisambigItems:(.*?$)')
        for line_num, line in enumerate(concept_file):
            line = line.strip().decode('utf-8')
            m = re_tag.match(line)
            assert m is not None, u'the {} line of "{}" is wrong!: {}'.format(line_num+1, self.concept_path, line)
            self.valid_word_concept[m.group(1)] = {
                u'Title': self.reduce_input(m.group(2)),
                u'RedirectTitle': self.reduce_input(m.group(3)),
                u'bPageErr': self.reduce_input(m.group(4)),
                u'DisambigItems': self.reduce_input(m.group(5), True),
                u'HatnoteLinkItems': self.reduce_input(m.group(6), True),
                u'HatnoteLinkDisambig': self.reduce_input(m.group(7), True),
                u'HatnoteLinkDisambigItems': self.reduce_input(m.group(8), True)
            }

    def reduce_input(self, text, bsplit=False):
        if text == u'None':
            return None
        if text == u'True':
            return True
        if bsplit:
            return text.split(u'\t')
        return text

    def reduce_output(self, text, bjoin=None):
        if text:
            if bjoin:
                return bjoin.join(text)
            return text
        return None

    def compute_similarity(self, model):
        """
        计算词语的相似度有四个维度，可信度依次递减：1 标题，重定向页；2 消岐页；3 hatnote指向的普通页或重定向页；4 hatnote指向的消岐页消岐页
        :return:
        """
        for w1, w2 in self.wordpairs:
            logger.info(u"compute similarity: {} | {}".format(w1, w2))
            # 获取两个词语的概念集合
            w1_concept = self.valid_word_concept[w1]
            w2_concept = self.valid_word_concept[w2]
            # 验证word是否在wiki中，bPageErr=False说明不在维基百科中
            bPageErr_w1 = w1_concept[u'bPageErr']
            bPageErr_w2 = w2_concept[u'bPageErr']
            if bPageErr_w1 == True and bPageErr_w2 == True:
                logger.warning(u'"{}" and "{}" doesn\'t exist in Wikipedia'.format(w1, w2))
            elif bPageErr_w1 == True and bPageErr_w2 != True:
                logger.warning(u'"{}" doesn\'t exist in Wikipedia'.format(w1))
            elif bPageErr_w2 == True and bPageErr_w1 != True:
                logger.warning(u'"{}" doesn\'t exist in Wikipedia'.format(w2))
            # 提取word在各个维度的概念
            w1_concept_list = self.get_concept_list(w1_concept)  # [ (标题，来源) ]
            w2_concept_list = self.get_concept_list(w2_concept)
            logger.info(u"{} concept_list : {}".format(w1, w1_concept_list))
            logger.info(u"{} concept_list : {}".format(w2, w2_concept_list))
            if not w1_concept_list and not w2_concept_list:
                logger.warning((u'{} | {}'.format(w1, w2)).encode('utf-8'))
            elif not w1_concept_list:
                logger.warning((u'{} | '.format(w1)).encode('utf-8'))
            elif not w2_concept_list:
                logger.warning((u' | {}'.format(w2)).encode('utf-8'))
            # 将两个词语的概念做笛卡尔集
            cartesianProductOfConcepts = [x for x in product(w1_concept_list, w2_concept_list)]  # [ ((标题1，来源), (标题2，来源)) ]
            source_list = []
            for w1_temp, w2_temp in cartesianProductOfConcepts:
                source_list.append((w1_temp[1], # 来源
                                    w2_temp[1], # 来源
                                    w1_temp[0], # 词1
                                    w2_temp[0], # 词2
                                    model.model.similarity(w1_temp[0], w2_temp[0]))) # 相似度计算结果

            self.result[(w1, w2)] = source_list


    def select_score_1(self):
        """
        分数选择策略，从已经计算出result的结果中选择结果，本策略说明如下：
        1. Title_Redirect 有结果的，选择最大值 * a
        2. DisambigItems 有结果的，选择最大值 * b
        3. HatnoteLinkItems 选择最大值 * c
        4. HatnoteLinkDisambig 选择最大值 * d
        5. HatnoteLinkDisambigItems 选择最大值 * e
        :return:
        """
        new_score = {}
        for w, scorce in self.result.iteritems():
            TRS = []
            HLI = []
            DI = []
            HLD = []
            HLDI = []
            for w1_s, w2_s, w1, w2, s in scorce:
                if w1_s in self.Title_Redirect and w2_s in self.Title_Redirect:
                    if TRS and TRS[4] < s:
                        TRS = [w1_s, w2_s, w1, w2, s]
                    elif not TRS:
                        TRS = [w1_s, w2_s, w1, w2, s]
                elif w1_s in self.DisambigItems and w2_s in self.DisambigItems:
                    if DI and DI[4] < s:
                        DI = [w1_s, w2_s, w1, w2, s]
                    elif not DI:
                        DI = [w1_s, w2_s, w1, w2, s]
                elif w1_s in self.HatnoteLinkItems and w2_s in self.HatnoteLinkItems:
                    if HLI and HLI[4] < s:
                        HLI = [w1_s, w2_s, w1, w2, s]
                    elif not HLI:
                        HLI = [w1_s, w2_s, w1, w2, s]
                elif w1_s in self.HatnoteLinkDisambig and w2_s in self.HatnoteLinkDisambig:
                    if HLD and HLD[4] < s:
                        HLD = [w1_s, w2_s, w1, w2, s]
                    elif not HLD:
                        HLD = [w1_s, w2_s, w1, w2, s]
                elif w1_s in self.HatnoteLinkDisambigItems and w2_s in self.HatnoteLinkDisambigItems:
                    if HLDI and HLDI[4] < s:
                        HLDI = [w1_s, w2_s, w1, w2, s]
                    elif not HLDI:
                        HLDI = [w1_s, w2_s, w1, w2, s]
                else:
                    raise ValueError, u"发现未见过的来源: {}, {}, {}, {}, {}".format(w1_s, w2_s, w1, w2, s)
            if TRS:
                new_score[w] = TRS
            elif DI:
                new_score[w] = DI
            elif HLI:
                new_score[w] = HLI
            elif HLD:
                new_score[w] = HLD
            elif HLDI:
                new_score[w] = HLDI
            else:
                #logger.error(u'{} | {} 没有计算结果，请检查程序.'.format(w[0], w[1]).encode('utf-8'))
                pass
        return new_score

    def add_weight_for_score(self, score, a, b, c):
        new_score = {}
        for word, (w1_s, w2_s, w1, w2, s) in score.iteritems():
            if w1_s in self.Title_Redirect and w2_s in self.Title_Redirect:
                new_score[word] = [w1_s, w2_s, w1, w2, s * a]
            elif w1_s in self.DisambigItems and w2_s in self.DisambigItems:
                new_score[word] = [w1_s, w2_s, w1, w2, s * b]
            #elif w1_s in self.HatnoteLinkItems and w2_s in self.HatnoteLinkItems:
            #    new_score[word] = [w1_s, w2_s, w1, w2, s * b]
            #elif w1_s in self.HatnoteLinkDisambig and w2_s in self.HatnoteLinkDisambig:
            #    new_score[word] = [w1_s, w2_s, w1, w2, s * c]
            elif w1_s in self.HatnoteLinkDisambigItems and w2_s in self.HatnoteLinkDisambigItems:
                new_score[word] = [w1_s, w2_s, w1, w2, s * c]
            else:
                raise ValueError, u"发现未见过的来源: {}, {}, {}, {}, {}".format(w1_s, w2_s, w1, w2, s)
        return new_score


    def spearman(self, score):
        humanscoreslist = []
        autorunscoreslist = []
        no_scorce_num = 0
        for k, v in self.dict_wordpair2humanscore.iteritems():
            if k in score:
                humanscoreslist.append(self.dict_wordpair2humanscore[k])
                autorunscoreslist.append(score[k][4])
            else:
                # 没有结果的词按照0.5处理
                humanscoreslist.append(self.dict_wordpair2humanscore[k])
                autorunscoreslist.append(0.5)
                #no_scorce_num += 1
        # max_scorce = max(autorunscoreslist)
        # min_scorce = min(autorunscoreslist)
        # ave = (max_scorce - min_scorce) / len(autorunscoreslist)
        # for k,v in enumerate(autorunscoreslist):
        #     if v == 0:
        #         autorunscoreslist[k] = ave
        [r, p] = spearmanr(humanscoreslist, autorunscoreslist)
        note = u'Spearman: nonNumber exist and be removed.{}/{}.'.format(no_scorce_num, len(self.wordpairs))
        return [r, p, note]


    def get_concept_list(self, concept):
        concept_list = []  # [ (title, source) ]
        if concept[u'Title']: concept_list.append((concept[u'Title'], u'Title'))
        if concept[u'RedirectTitle']: concept_list.append((concept[u'RedirectTitle'], u'RedirectTitle'))
        if concept[u'DisambigItems']:
            concept_list.extend([x for x in product(concept[u'DisambigItems'], [u'DisambigItems'])])
        if concept[u'HatnoteLinkItems']:
            concept_list.extend([x for x in product(concept[u'HatnoteLinkItems'], [u'HatnoteLinkItems'])])
        if concept[u'HatnoteLinkDisambig']:
            concept_list.extend([x for x in product(concept[u'HatnoteLinkDisambig'], [u'HatnoteLinkDisambig'])])
        if concept[u'HatnoteLinkDisambigItems']:
            concept_list.extend(
                [x for x in product(concept[u'HatnoteLinkDisambigItems'], [u'HatnoteLinkDisambigItems'])])
        return concept_list



class ConceptModel:

    def __init__(self, model_path, is_cce_hy=False):
        self.model_path = model_path
        self.model = None
        self.id2StandardWikititle = {}
        self.wikititle2Id = {}
        self.is_cce_hy = is_cce_hy

        # self.load_word2vec_model()
        # self.read_map_id_and_lowercompact_or_standard_title()


    def load_word2vec_model(self):
        """
         the path of model. note: the path must be end with .gen  .bin  .vec
         the mdoel is generted by gensimword2vec/TrainWordEmbeddings.py.
         .gen is corresponding with the formation of gensim, which can continue training with more data
         .bin is corresponding with the formation of google word2vec, which is C binary format
         .vec is corresponding with the formation of google word2vec, which is C text format
        :return:
        """
        assert os.path.isfile(self.model_path)
        print('load model')
        lastpart = self.model_path[self.model_path.rfind(u'.'):]
        assert lastpart == u'.vec' or lastpart == u'.gen' or lastpart == u'.bin', u'modelpath is:{} ,which is wrong!'.format(self.model_path)
        if lastpart == u'.vec':
            self.model = KeyedVectors.load_word2vec_format(self.model_path, binary=False, encoding='utf-8')
        elif lastpart == u'.bin':
            self.model = KeyedVectors.load_word2vec_format(self.model_path, binary=True)
        elif lastpart == u'.gen':
            self.model = Word2Vec.load(self.model_path)
        else:
            assert False, u'need to update source code'

    def get_vec_index(self, word):
        """
        获取word在词向量中的索引值 id_StandWikiTitle
        :param word:
        :return:
        """
        id_inWiki = self.look_for_id_of_wikititle(word)  # 对Antonín Dvořák查找到其Id: 76572。涉及到读入映射表的过程
        if id_inWiki != globalnonExistingId:
            # get this id's standard wiki title
            if self.is_cce_hy:
                id_concepttext = unicode(id_inWiki)
            else:
                standardTitle = self.id2StandardWikititle.get(id_inWiki)
                newlist_wikititle = standardTitle.split(u' ')
                id_inWiki = unicode(id_inWiki)
                newlist_wikititle.insert(0, id_inWiki)  # [ 23040, political , philosophy ]
                id_concepttext = u'_'.join(newlist_wikititle)
            if id_concepttext in self.model.index2word:
                return id_concepttext
        return None

    def look_for_id_of_wikititle(self, title):
        assert self.id2StandardWikititle and self.wikititle2Id, '映射表为空，请检查'
        #lower and compact the title
        lc_title = title
        lc_title = lc_title.replace(u' ',u'')
        lc_title = lc_title.lower()

        if self.wikititle2Id.has_key(lc_title):# use LowerCompactTitle to look
            return self.wikititle2Id.get(lc_title)
        else:#use StandardTitle to look
            firstchar = title[0:1]
            otherchars = title[1:]
            firstchar = firstchar.upper()
            titlefirstcharupper = u'{}{}'.format(firstchar,otherchars)
            if not self.wikititle2Id.has_key(titlefirstcharupper):
                title_CapitalFirstWord = title.capitalize()
                if not self.wikititle2Id.has_key(title_CapitalFirstWord):
                    #print title_CapitalFirstWord
                    title_Titlelize = title.title()
                    if not self.wikititle2Id.has_key(title_Titlelize):
                        title_Titlelize = title_Titlelize.replace(u' Of ', u' of ')
                        title_Titlelize = title_Titlelize.replace(u' The ', u' the ')
                        title_Titlelize = title_Titlelize.replace(u' In ', u' in ')
                        title_Titlelize = title_Titlelize.replace(u' On ', u' on ')
                        title_Titlelize = title_Titlelize.replace(u' To ', u' to ')
                        title_Titlelize = title_Titlelize.replace(u' At ', u' at ')
                        if not self.wikititle2Id.has_key(title_Titlelize):
                             ret = globalnonExistingId # 99999999 means fail to find its id
                        else:
                            ret = self.wikititle2Id[title_Titlelize]
                    else:
                        ret = self.wikititle2Id[title_Titlelize]
                else:
                    ret = self.wikititle2Id[title_CapitalFirstWord]
            else:
                ret = self.wikititle2Id[titlefirstcharupper]

        return ret

    def read_map_id_and_lowercompact_or_standard_title(self):
        if len(self.id2StandardWikititle) == 0 and len(self.wikititle2Id) == 0:
            print('load pickle')
            path_pickle = '/home/zero/workspace/wiki/Datas/WikiExtracter/enwiki-20161001-pages-articles-multistream-index_IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle-pickle.pkl'
            assert os.path.isfile(path_pickle), '映射表文件不存在，请检查'
            pklfile = file(path_pickle, 'rb')
            # pickle的变量加载load顺序必须要与保存dump顺序完全一致
            self.id2StandardWikititle = pickle.load(pklfile)
            self.wikititle2Id = pickle.load(pklfile)
            pklfile.close()