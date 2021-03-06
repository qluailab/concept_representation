# -*- coding: utf-8 -*-

import os
import sys
import time
import argparse
import logging
from enum import Enum
from openpyxl.reader.excel import load_workbook
import numpy
from gensim.models import Word2Vec, KeyedVectors
from gensim import matutils
import copy
from pattern.en import parsetree
from nltk.corpus import stopwords
import re
import fileinput
import itertools
import cPickle as pickle


# DataSet = Enum('DataSet', ('WordSim353', 'SemEval2017Task2En', 'RG65', 'MEN3000', 'MTURK771', 'RW2034', 'WS353SIM', 'WS353REL', 'MC30'))
#DataSet = Enum('WordSim353', 'SemEval2017Task2En', 'RG65', 'MEN3000', 'MTURK771', 'RW2034', 'WS353SIM', 'WS353REL')
SimilarityMode = Enum('SimilarityMode', ('OnlyUseWordForm', 'OnlyUseWordConceptFromWikibyConceptVec', 'UseWordFormAndWordConcept', 'OnlyUseWordConceptFromNode2vec', 'OnlyUserCombineWikiConceptNode2vec'))
PreprocFormsOfWordPair = Enum('PreprocFormsOfWordPair', ('OriginalToken', 'OnlyLowercase', 'LemmaLowercase'))
MultiwordsvecBySumSinglewordvec = Enum('MultiwordsvecBySumSinglewordvec', ('DontSumAnywords', 'SumAllwords', 'SumAllwordsExceptforStopwords',\
                                                                           'OnlySumContentwordscontainedinWN30'))
LookupConceptOfWordbyWiki = Enum('LookupConceptOfWordbyWiki',('Normal', 'Normal_Redirect', 'Normal_Redirect_Disambig', 'Normal_Redirect_HatenoteLinkItems',
                                                              'Normal_Redirect_Disambig_HatenoteLinkItems', 'Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems'))
SelectFromScoreSetofConceptPair = Enum('SelectFromScoreSetofConceptPair', ('MAX_ALL', 'MAX_TitleRedirectTitle', 'MAX_TitleRedirectTitleDisambigItems', 'PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_TitleRedirectTitleDisambigItemsHatnoteLinkItemsHatnoteLinkDisambigItems'))
gl_ConceptSourceFromWikipedia = [u'Title', u'RedirectTitle', u'DisambigItems', u'HatnoteLinkItems', u'HatnoteLinkDisambigItems']

base_path = os.path.split(os.path.realpath(__file__))[0]

# gl_wordsim353path = u'{}/data/wordsim353score.xlsx'.format(base_path)
# gl_wordsemeval2017task2enpath = u'{}/data/SemEval2017Task2Enscore.xlsx'.format(base_path)
# gl_rg65path = u'{}/data/RG65score.xlsx'.format(base_path)
# gl_men3000path = u'{}/data/MEN3000score.xlsx'.format(base_path)
# gl_mturk771path = u'{}/data/MTurk771score.xlsx'.format(base_path)
# gl_rw2034path = u'{}/data/RW2034score.xlsx'.format(base_path)
# gl_ws353simpath = u'{}/data/WS353-SIM-score.xlsx'.format(base_path)
# gl_ws353relpath = u'{}/data/WS353-REL-score.xlsx'.format(base_path)
# gl_mc30path = u'{}/data/MC30score.xlsx'.format(base_path)
# gl_allscorecomparepath = u'{}/data/AllscoreCompare.xlsx'.format(base_path)
gl_modelembedding1 = None # the model of word2vec, load from embeddingFile1
gl_modelembedding2 = None # the model of word2vec, load from embeddingFile2
gl_Dict_stopwords = {}  # call init_glDictStopword()
gl_Dict_contentwords = {} # call init_glDictContentwords()

# gl_conceptwordsim353path = u'{}/data/dataset/word353/combined.concept.txt'.format(base_path)
# gl_conceptsemeval2017task2enpath = u'{}/data/dataset/SemEval 2017 task2/en.test.data.concept.txt'.format(base_path)
# gl_conceptrg65path = u'{}/data/dataset/RG65.concept.txt'.format(base_path)
# gl_conceptmen3000path = u'{}/data/dataset/MEN3000.concept.txt'.format(base_path)
# gl_conceptmturk771path = u'{}/data/dataset/MTurk771.concept.txt'.format(base_path)
# gl_conceptrw2034path = u'{}/data/dataset/RW2034.concept.txt'.format(base_path)
# gl_conceptws353simpath = u'{}/data/dataset/WS353SIM.concept.txt'.format(base_path)
# gl_conceptws353relpath = u'{}/data/dataset/WS353REL.concept.txt'.format(base_path)
# gl_conceptmc30path = u'{}/data/dataset/MC30.concept.txt'.format(base_path)

gl_conceptpaht = u'{}/data/concept.txt'.format(base_path)

gl_cache = './data/cache/'

globalDictId2StandardWikititle = {} # Id --->  Standard Wikipedia Title
globalDictWikititle2Id = {} # Lowerconpacttitle or Standardtitle ---> Id
globalnonExistingId = 999999999
gl_simScoreRepresentOneObjectMissing = -99 # means that there is a missing concept in (c1,c2) by vector model


gl_flagPreprocFormsOfWordPair = PreprocFormsOfWordPair.LemmaLowercase#OriginalToken OnlyLowercase LemmaLowercase#词对在进行比较之前，
                                                                    # 是直接使用原样词；还是将它们改为小写；还是将它们进行词形还原并小写

gl_flagMultiwordsvecBySumsinglewordvec = MultiwordsvecBySumSinglewordvec.DontSumAnywords
                                            # DontSumAnywords SumAllwords SumAllwordsExceptforStopwords OnlySumContentwordscontainedinWN30
                                            # #getwordvectorForPhraseFromModelOnlybyWordForm()函数中，
                                            # 对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。

#这个flagLookupConceptOfWordbyWiki参数的使用主要就是控制将 事先准备的由wikipedia查到的词语的候选概念，选择哪些读进来以参与 概念相似度的计算。
#其作用感觉已被flagtoSelectFromScoreSetofConceptPair所取代。
#后者，比前者要粗略一些，不如前者划分细致。但是对于概念相似度计算感觉也已经足够。
#除非有特别细致要求，flagLookupConceptOfWordbyWiki的参数就固定为最大范围就可以。
#这个变量打算直接固定为最大范围了。（除非非常确定就只使用某个小范围就足够有效，不要改动了）!!!!!!!!!!!!!!
gl_flagLookupConceptOfWordbyWiki = LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
assert gl_flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
                                            #Normal Normal_Redirect Normal_Redirect_Disambig Normal_Redirect_HatenoteLinkItems
                                            # Normal_Redirect_Disambig_HatenoteLinkItems
                                            #Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
                                            #当利用Wikipedia查找一个Word所对应的Concept时的查找层次级别. 看程序上方的详细说明

#当两个词的概念集合都找到，并计算完概念对的相似度后，会有很多值，从中选哪个值返回，由这个参数决定
gl_flagtoSelectFromScoreSetofConceptPair = SelectFromScoreSetofConceptPair.PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_TitleRedirectTitleDisambigItemsHatnoteLinkItemsHatnoteLinkDisambigItems
                                            # MAX_ALL MAX_TitleRedirectTitle MAX_TitleRedirectTitleDisambigItems
                                            # PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_TitleRedirectTitleDisambigItemsHatnoteLinkItemsHatnoteLinkDisambigItems
                                            # 看程序上方的详细说明
# gl_flagtoSelectFromScoreSetofConceptPair = SelectFromScoreSetofConceptPair.MAX_ALL

#'s_gravenhage%1:15:00:: 08970180 1 0
reTag_wn31dict = re.compile(r'^(.*?)%\d:.*$')


if sys.platform == "win32":
    # On Windows, the best timer is time.clock()
    default_timer = time.clock
else:
    # On most other platforms the best timer is time.time()
    default_timer = time.time


def get_time_info():
    return time.strftime('%m%d%H%M%S',time.localtime(time.time()))


#=============================================================================
#  Description:     获取当前位置的行号和函数名
#  Version:         1.0
#  LastChange:      2010-12-17 01:19:19
#  History:
#=============================================================================

def get_cur_info():
    """Return the frame object for the caller's stack frame."""
    try:
        raise Exception
    except:
        f = sys.exc_info()[2].tb_frame.f_back
    return (f.f_code.co_name, f.f_lineno)





def init_glDictStopword():
    tt = stopwords.words('english')
    for w in tt:
        gl_Dict_stopwords[w]=u''
    assert len(gl_Dict_stopwords) > 0
    # following is self-defined
    gl_Dict_stopwords[u'\'s']=u''
    gl_Dict_stopwords[u'n\'t']=u''


def init_glDictContentwords():
    wn31path = u'./data/win3.1.dict.index.sense'
    assert os.path.isfile(wn31path)==True, u'{}: {} doesn\'t exist'.format(get_cur_info(), wn31path)
    fh = open(wn31path, 'r')
    for line in fh.readlines():
        line = line.decode('utf-8')
        m = reTag_wn31dict.match(line)
        if m == None:
            print('may be a error!!!!!!!!!!!!!')
            pass
        else:
            word = m.group(1)
            word = word.replace(u'_',u' ')
            if gl_Dict_contentwords.has_key(word) == False:
                gl_Dict_contentwords[word]=u''
    print('Content words dictionary has been loaded({})! {}'.format(len(gl_Dict_contentwords), wn31path))

class DumpRunTime:
    '''
    打印执行时间
    with DumpRunTime('build_link'):
        build_link()
    '''

    def __init__(self, tag):
        self.tag = tag
        self.startTime = default_timer()

    def __enter__(self):
        return self  # 可以返回不同的对象

    def __exit__(self, exc_type, exc_value, exc_tb):
        if exc_tb is None:
            use_time = default_timer() - self.startTime
            logging.debug('[%s] run time is %.4fs', self.tag, use_time)
        else:
            logging.error('[Exit %s]: Exited with %s, %s.', self.tag, str(exc_type), str(exc_value))


class Timer(object):
    """
    计时器，对于需要计时的代码进行with操作：
    with Timer() as timer:
        ...
        ...
    print(timer.cost)
    ...
    """
    def __init__(self, start=None):
        self.start = start if start is not None else time.time()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop = time.time()
        self.cost = self.stop - self.start
        return exc_type is None


def getStrOfdictWordpair2Score(dictWordpair2Score, list_orderWordpair = None):
    '''
    return the String form of dictWordpair2Score
    :param dictWordpair2Score: 
    :return: 
    '''
    ret = u''.encode('utf-8')
    if list_orderWordpair == None:
        for wordpair in dictWordpair2Score:
            line = u'{}\t{}\n'.format(wordpair, dictWordpair2Score.get(wordpair))
            line = line.encode('utf-8')
            ret += line
    else:
        for i in range(0, len(list_orderWordpair)):
            wordpair = list_orderWordpair[i]
            line = u'{}\t{}\n'.format(wordpair, dictWordpair2Score.get(wordpair))
            line = line.encode('utf-8')
            ret += line
    #print ret
    return ret


def loadModelforWord2Vec(modelpath):
    '''
    :param modelpath: the path of model. note: the path must be end with .gen  .bin  .vec
     the mdoel is generted by gensimword2vec/TrainWordEmbeddings.py. 
     .gen is corresponding with the formation of gensim, which can continue training with more data
     .bin is corresponding with the formation of google word2vec, which is C binary format
     .vec is corresponding with the formation of google word2vec, which is C text format
    :return: 
    '''
    model = None
    assert os.path.isfile(modelpath)
    lastpart = modelpath[modelpath.rfind(u'.'):]
    assert lastpart == u'.vec' or lastpart == u'.gen' or lastpart == u'.bin', u'{}: modelpath is:{} ,which is wrong!'.format(get_cur_info(), modelpath)
    if lastpart == u'.vec':
        #model = Word2Vec.load_word2vec_format(modelpath, binary=False, encoding='utf-8')
        model = KeyedVectors.load_word2vec_format(modelpath, binary=False, encoding='utf-8')
    elif lastpart == u'.bin':
        #model = Word2Vec.load_word2vec_format(modelpath, binary=True)
        model = KeyedVectors.load_word2vec_format(modelpath, binary=True)
    elif lastpart == u'.gen':
        model = Word2Vec.load(modelpath)
    else:
        assert False, u'{}: need to update source code'.format(get_cur_info())
    return model


def lines_from(input):
    for line in input:
        line = line.decode('utf-8')
        line = line.strip('\n\r')
        yield line


def str2list(text):
    #text = 'James Prescott Joule	Joule heating	Joule effect	Joule (crater)	Joule (programming language)	Joule (surname)	Optimal Energy Joule'
    #print text
    ret = []
    for t in text.split(u'\t'):
        ret.append(t)
    if len(ret)==0:
        return None
    else:
        return ret



def readLookupFileWord2ConceptbyWiki(filepath):
    '''
    将使用QueryWikiInfoForWordPairFile.py得到的各个词语所对应的Wiki概念信息，读入内存，以备后面相似度计算使用。并显示一个初步的统计信息。
    :param filepath: 
    :return: 
    '''
    assert os.path.isfile(filepath), u'{}: {} is not a real existing file!'.format(get_cur_info(), filepath)

    DictTitle2ConceptInfo = {}

    conceptfile = fileinput.FileInput(filepath)
    reTag = re.compile(r'^Title:(.*?)    bNormal:(.*?)    bDisambig:(.*?)    bRedirect:(.*?)    RedirectTitle:(.*?)    bPageErr:(.*?)    bHttpTimeOut:(.*?)    DisambigItems:(.*?)    HatnoteLinkItems:(.*?)    HatnoteLinkDisambig:(.*?)    HatnoteLinkDisambigItems:(.*?)$')
    zongshu = 0
    for line in lines_from(conceptfile):
        m = reTag.match(line)
        if m == None and len(line)>=1:
            assert False, u'{}: the last line of "{}" is wrong!: {}'.format(get_cur_info(), filepath, line)
        if m == None:
            continue
        #从文件中读出，存入变量
        zongshu += 1
        Title = m.group(1)

        if m.group(2)==u'True':
            bNormal = True
        elif m.group(2)==u'False':
            bNormal = False
        else:
            assert False, u'{}: {} need to update handle code'.format(get_cur_info(), m.group(2))

        if m.group(3)==u'True':
            bDisambig = True
        elif m.group(3)==u'None':
            bDisambig = None
        else:
            assert False, u'{}: {} need to update handle code'.format(get_cur_info(), m.group(3))

        if m.group(4)==u'None':
            bRedirect = None
        elif m.group(4)==u'True':
            bRedirect = True
        else:
            assert False, u'{}: {} need to update handle code'.format(get_cur_info(), m.group(4))

        if m.group(5)==u'None':
            RedirectTitle = None
        else:
            RedirectTitle = m.group(5)

        if m.group(6)==u'None':
            bPageErr = None
        elif m.group(6)==u'True':
            bPageErr = True
        else:
            assert False, u'{}: {} need to update handle code'.format(get_cur_info(), m.group(6))

        if m.group(7)==u'None':
            bHttpTimeOut = None
        elif m.group(7)==u'True':
            bHttpTimeOut = True
        else:
            assert False, u'{}: {} need to update handle code'.format(get_cur_info(), m.group(7))

        if m.group(8)==u'None':
            DisambigItems = None
        else:
            DisambigItems = str2list(m.group(8))
            #对dolloar所含“V	T	E	”的特殊处理
            for vte in [u'V', u'T', u'E']:
                if vte in DisambigItems:
                    DisambigItems.remove(vte)

        if m.group(9)==u'None':
            HatnoteLinkItems = None
        else:
            HatnoteLinkItems = str2list(m.group(9))

        if m.group(10)==u'None':
            HatnoteLinkDisambig = None
        else:
            HatnoteLinkDisambig = str2list(m.group(10))

        if m.group(11)==u'None':
            HatnoteLinkDisambigItems = None
        else:
            HatnoteLinkDisambigItems = str2list(m.group(11))
            # 对dolloar所含“V	T	E	”的特殊处理
            for vte in [u'V', u'T', u'E']:
                if vte in HatnoteLinkDisambigItems:
                    HatnoteLinkDisambigItems.remove(vte)

        #将信息存入词典，以备以后使用
        if DictTitle2ConceptInfo.has_key(Title):
            conceptinfo = DictTitle2ConceptInfo[Title]
            conceptinfo[u'CountInCurDataset'] += 1
        else:
            conceptinfo = {}
            conceptinfo[u'Title'] = Title
            conceptinfo[u'bNormal'] = bNormal
            conceptinfo[u'bDisambig'] = bDisambig
            conceptinfo[u'bRedirect'] = bRedirect
            conceptinfo[u'RedirectTitle'] = RedirectTitle
            conceptinfo[u'bPageErr'] = bPageErr
            conceptinfo[u'bHttpTimeOut'] = bHttpTimeOut
            conceptinfo[u'DisambigItems'] = DisambigItems
            conceptinfo[u'HatnoteLinkItems'] = HatnoteLinkItems
            conceptinfo[u'HatnoteLinkDisambig'] = HatnoteLinkDisambig
            conceptinfo[u'HatnoteLinkDisambigItems'] = HatnoteLinkDisambigItems

            conceptinfo[u'CountInCurDataset'] = 1
            DictTitle2ConceptInfo[Title] = conceptinfo

    conceptfile.close()
    #由wiki info的文件文本统计 各种情况的频次信息
    staticsDictTitle2ConceptInfo(DictTitle2ConceptInfo)
    return DictTitle2ConceptInfo






def staticsDictTitle2ConceptInfo(DictTitle2ConceptInfo):
    # 统计各种情形(歧义页、重定向、不存在、访问超时、存在HatnoteLinkItems或HatnoteLinkDisambig)的Title的数量：

    bNormaltitles = {}
    bDisambigtitles = {}
    bRedirecttitles = {}
    bPageErrtitles = {}
    bHttpTimeOuttitles = {}
    bHatnoteLinkItemstitles = {}
    bHatnoteLinkDisambigtitles = {}

    bRedirectAndDisambigtitles = {}
    bRedirectAndPageErrtitles = {}
    bRedirectAndHttpTimeOuttitles = {}


    zongshu = 0
    for Title in DictTitle2ConceptInfo.keys():
        zongshu += DictTitle2ConceptInfo[Title][u'CountInCurDataset']

        if DictTitle2ConceptInfo[Title][u'bNormal'] == True:
            if bNormaltitles.has_key(Title):
                assert False, u'{}: {} should not exist in bNormaltitles!'.format(get_cur_info(), Title)
            else:
                bNormaltitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'bDisambig'] == True:
            if bDisambigtitles.has_key(Title):
                assert False
            else:
                bDisambigtitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'bRedirect'] == True:
            if bRedirecttitles.has_key(Title):
                assert False
            else:
                bRedirecttitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
                #统计那些是重定向页面并且又为歧义页或不存在的页面
                if DictTitle2ConceptInfo[Title][u'bDisambig'] == True:
                    if bRedirectAndDisambigtitles.has_key(Title):
                        assert False
                    else:
                        bRedirectAndDisambigtitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
                if DictTitle2ConceptInfo[Title][u'bPageErr'] == True:
                    if bRedirectAndPageErrtitles.has_key(Title):
                        assert False
                    else:
                        bRedirectAndPageErrtitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
                if DictTitle2ConceptInfo[Title][u'bHttpTimeOut'] == True:
                    if bRedirectAndHttpTimeOuttitles.has_key(Title):
                        assert False
                    else:
                        bRedirectAndHttpTimeOuttitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'bPageErr'] == True:
            if bPageErrtitles.has_key(Title):
                assert False
            else:
                bPageErrtitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'bHttpTimeOut'] == True:
            if bHttpTimeOuttitles.has_key(Title):
                assert False
            else:
                bHttpTimeOuttitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'HatnoteLinkItems'] != None:
            if bHatnoteLinkItemstitles.has_key(Title):
                assert False
            else:
                bHatnoteLinkItemstitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']
        if DictTitle2ConceptInfo[Title][u'HatnoteLinkDisambig'] != None:
            if bHatnoteLinkDisambigtitles.has_key(Title):
                assert False
            else:
                bHatnoteLinkDisambigtitles[Title] = DictTitle2ConceptInfo[Title][u'CountInCurDataset']


    # 输出统计信息
    logging.info('Statistical info of all compared words:')
    logging.info("\tCount of word tokens:{}".format(zongshu))
    logging.info("\tCount of word types:{}".format(len(DictTitle2ConceptInfo.keys())))
    logging.debug("\tas follows:{}".format(u'  ,  '.join(DictTitle2ConceptInfo.keys()).encode('utf-8')))
    sum = 0
    for key in DictTitle2ConceptInfo.keys():
        sum += DictTitle2ConceptInfo[key][u'CountInCurDataset']
    assert zongshu == sum
    logging.info("\tFrequency:%d  %.2f%%" % (sum, sum/sum*100))
    tdict = {}
    for t in DictTitle2ConceptInfo.keys():
        tdict[t] = DictTitle2ConceptInfo[t][u'CountInCurDataset']
    logging.debug("\tOrdered by frequency:{}".format(sorted(tdict.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)))


    #logging.info('bNormal为True的待比较词统计信息:')
    logging.info('Statistical info of compared words "bNormal=True":')
    freqNormaltitles = printStatisInfo(bNormaltitles, zongshu)

    # logging.info('bRedirect为True的待比较词统计信息:')
    logging.info('Statistical info of compared words "bRedirect=True":')
    freqRedirecttitles = printStatisInfo(bRedirecttitles, zongshu)

    #logging.info('bPageErr为True的待比较词统计信息:')
    logging.info('Statistical info of compared words "bPageErr=True":')
    freqPageErrtitles = printStatisInfo(bPageErrtitles, zongshu)

    #logging.info('bHttpTimeOut为True的待比较词统计信息:')
    logging.info('Statistical info of compared words "bHttpTimeOut=True":')
    freqHttpTimeOuttitles = printStatisInfo(bHttpTimeOuttitles, zongshu)
    if freqHttpTimeOuttitles > 0:
        assert False, u'{}: the number of HttpTimeOuttitle is:{}. You need to requery wiki for them!'.format(get_cur_info(), freqHttpTimeOuttitles)

    #logging.info('bDisambig为True的待比较词统计信息:')
    logging.info('Statistical info of compared words "bDisambig=True":')
    freqDisambigtitles = printStatisInfo(bDisambigtitles, zongshu)

    #logging.info('HatnoteLinkItems不为None的待比较词统计信息:')
    logging.info('Statistical info of compared words "HatnoteLinkItems!=None":')
    freqHatnoteLinkItemstitles = printStatisInfo(bHatnoteLinkItemstitles, zongshu)

    #logging.info('HatnoteLinkDisambig不为None的待比较词统计信息:')
    logging.info('Statistical info of compared words "HatnoteLinkDisambig!=None":')
    freqHatnoteLinkDisambigtitles = printStatisInfo(bHatnoteLinkDisambigtitles, zongshu)

    logging.info('Statistical info of compared words "bRedirect=True and bDisambig=True":')
    freqRedirectAndDisambigtitles = printStatisInfo(bRedirectAndDisambigtitles, zongshu)

    logging.info('Statistical info of compared words "bRedirect=True and bPageErr=True":')
    freqRedirectAndPageErrtitles = printStatisInfo(bRedirectAndPageErrtitles, zongshu)

    logging.info('Statistical info of compared words "bRedirect=True and bHttpTimeOut=True":')
    freqRedirectAndHttpTimeOuttitles = printStatisInfo(bRedirectAndHttpTimeOuttitles, zongshu)

    assert (freqNormaltitles + freqRedirecttitles + freqPageErrtitles + freqHttpTimeOuttitles + freqDisambigtitles) - (freqRedirectAndDisambigtitles + freqRedirectAndPageErrtitles + freqRedirectAndHttpTimeOuttitles)  \
           == zongshu, u'{}: please check the statistical function!'

    return




def printStatisInfo(tdict, zongshu):
    #logging.info("\t总词形数：{}".format(len(tdict.keys())))
    logging.info("\tCount of word types:{}".format(len(tdict.keys())))
    if len(tdict.keys())==0:
        return 0
    #logging.info("\t分别为:{}".format(u'  ,  '.join(tdict.keys()).encode('utf-8')))
    logging.info("\tas follows:{}".format(u'  ,  '.join(tdict.keys()).encode('utf-8')))
    sum = 0
    for key in tdict.keys():
        sum += tdict[key]
    #logging.info("\t出现频次总数为:%d  %.2f%%" % (sum, (sum*1.0)/zongshu*100))
    logging.info("\tFrequency:%d  %.2f%%" % (sum, (sum * 1.0) / zongshu * 100))
    #logging.info("\t按频次降序排列：{}".format(sorted(tdict.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)))
    logging.info("\tOrdered by frequency:{}".format(sorted(tdict.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)))
    return sum

def getLemmasOfText(text):
    ret = []
    s = parsetree(text, relations = False, lemmata = True)
    for sen in s:
        ret.append(sen.lemmata)
    return ret

def isContentword(s):
    if len(gl_Dict_contentwords) == 0:
        init_glDictContentwords()
    assert len(gl_Dict_contentwords) > 0
    if gl_Dict_contentwords.get(s)!=None:
        return True
    else:
        return False

def isStopword(s):
    if len(gl_Dict_stopwords) == 0:
        init_glDictStopword()
    assert len(gl_Dict_stopwords) > 0
    if gl_Dict_stopwords.get(s)!=None:
        return True
    else:
        return False


def getChildwordOfPhrase(phrase, flagMultiwordsvecBySumsinglewordvec):
    ret = []
    if flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.DontSumAnywords:
        ret.append(phrase)
    elif flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.SumAllwords:
        words = phrase.split(u' ')
        ret = words
    elif flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.SumAllwordsExceptforStopwords:
        words = phrase.split(u' ')
        for w in words:
            if isStopword(w.lower()) == True:
                continue
            else:
                ret.append(w)
    elif flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.OnlySumContentwordscontainedinWN30:
        words = phrase.split(u' ')
        for w in words:
            if isContentword(w) == True:
                ret.append(w)
    else:
        assert False, u'{} need to update source code'.format(get_cur_info())
    return ret


def getconceptvectorForConceptFromModelOnlyUseWordConceptFromWikibyConceptVec(concepttext, model, flagMultiwordsvecBySumsinglewordvec, similarityMode=None):
    '''
    由model取得concepttext所对应的 概念向量
    
    :param concepttext: 概念的文本
    :param model: 使用的词向量文件
    :param flagMultiwordsvecBySumsinglewordvec: 对于比较长的概念文件，如果其在词向量中并不存在，如何利用它所包含的子概念来合成其整体概念
    :return: 概念文本所对应的概念向量.
    
    
    flagMultiwordsvecBySumSinglewordvec
    对于无法直接查得概念向量的concepttext， 如何利用它所包含的子概念来合成其整体概念
    1)'DontSumAnywords' 这个要当于对concepttext直接做放弃处理，不会对子概念作累加等处理去求．如果找不到，返回None。
    
    '''
    #目前只做了对这种最简单的情况的处理
    assert flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.DontSumAnywords, \
        u'{}: need to update source code to handle: {}'.format(get_cur_info(), flagMultiwordsvecBySumsinglewordvec)

    retvector = None
    logging.debug('flagMultiwordsvecBySumsinglewordvec: {}'.format(flagMultiwordsvecBySumsinglewordvec))

    if flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.DontSumAnywords:
        #查概念所对应的词义编号
        id_inWiki = LookforIDofWikititle(concepttext)#对Antonín Dvořák查找到其Id: 76572。涉及到读入映射表的过程
        if id_inWiki != globalnonExistingId:
            if similarityMode == SimilarityMode.OnlyUseWordConceptFromNode2vec:
                # zytcode node2vec生成的词向量没有id
                id_inWiki = unicode(id_inWiki)
                if id_inWiki in model.vocab:
                    retvector = model.vectors[model.vocab[id_inWiki].index]
                    # retvector = model.word_vec(id_inWiki)
            else:
                ###这个过程与ProWikiParsedOutToTrainCorpus.py中的ParseOneLine2ConceptLemmaOrConcept()是一样的
                # 对于特殊字符,以bahá'í和brønsted为例(用 人民 and 特殊　也验证了一下．),已验证下列代码。可以将它们从excel中作为词对读出，而后可以在词向量文件中查找到它们对应的词向量
                # get this id's standard wiki title
                standardTitle = globalDictId2StandardWikititle.get(id_inWiki)
                newlist_wikititle = standardTitle.split(u' ')
                # get the form:  id , words in wikititle
                #print '求wikipedia中的id的代码暂时注释掉!!!!!! {}'.format(get_cur_info())
                id_inWiki = unicode(id_inWiki)  # convert to string, then to unicode
                newlist_wikititle.insert(0, id_inWiki)  # [ 23040, political , philosophy ]
                # get the form: 20340_political_philosophy
                id_concepttext = u'_'.join(newlist_wikititle)
                ###
                # get concept vector from model

                if id_concepttext in model.index2word:
                    retvector = model.wv[id_concepttext]

    else:
        assert False, u'{}: need to update source code for: {}'.format(get_cur_info(), flagMultiwordsvecBySumsinglewordvec)

    return retvector




def getwordvectorForPhraseFromModelOnlybyWordForm(phrase, model, flagMultiwordsvecBySumsinglewordvec):
    '''
    由model取得短语（词组）或单词所对应的向量
    如果失败，返回None

    phrase为短语或单词
    model为词向量模型
    flagMultiwordsvecBySumsinglewordvec对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。

    flagMultiwordsvecBySumSinglewordvec
    对于无法直接查得词向量的multiwords短语词组，如何由其所包含的word来求得其词向量
    1)'DontSumAnywords' 这个要当于对multiwords直接做放弃处理，不会累加单个词的vector去求．
    2)'SumAllwords'　将multiwords中的每个词，无论是实词还是停用词，把它们的词向量都加起来．
    3)'SumAllwordsExceptforStopwords'将multiwords中的除停用词之外的词，词向量累加起来
    4)'OnlySumContentwordscontainedinWN30'只把multiwords中的实词的词向量累加起来,

    :param phrase: 
    :param model: 
    :return: 
    '''
    retvector = None
    retfoundwords = []

    # 对phrase中的单词进行过滤，看哪些是需要保留下来的（将用于累加得到的phrase向量的）
    satisfiedChildwords = getChildwordOfPhrase(phrase, flagMultiwordsvecBySumsinglewordvec)
    # 累加向量（当不允许累加时，satisfiedChildwords返回的是整个phrase，也就相当于直接以phrase为key去查值而不累加了）
    validcount = 0
    for w in satisfiedChildwords:
        if model.__contains__(w):
            retfoundwords.append(w)
            validcount += 1
            if retvector is None:
                retvector = model[w]
            else:
                retvector += model[w]

    retfoundwords = u' '.join(retfoundwords)
    if retfoundwords != phrase:
        logging.info('"%s" actually valid contained words in model are: "%s"    %.2f%%', phrase, retfoundwords,
                     validcount * 1.0 / len(phrase.split(u' ')) * 100)

    return retvector




def LookforIDofWikititle(title):
    '''
    该函数由ProcWikiParsedOutToTrainCorpus.py拷贝过来
    :param title:  the title that need to know its ID in wikipedia Dump index
    :return: its id
    '''


    #read Map Table(Id --- WikiTitle)
    #print 'for test {}'.format(get_cur_info())
    #if len(DictId2Wikititle) == 0 or len(DictWikititle2Id) == 0:
    #    assert ReadMapWikiIdAndTitle()==True, "{}: Failed to ReadMapWikiIdAndTitle()".format(get_cur_info())
        #AnalysisMapWikiIdAndTitle()


    # print 'for test {}'.format(get_cur_info())
    if len(globalDictId2StandardWikititle) == 0 or len(globalDictWikititle2Id) == 0:
        assert ReadMapIdAndLowercompactOrStandardTitle()==True, '{}: Failed to ReadMapIdAndLowercompactOrStandardTitle()'.format(get_cur_info())

    key = u'{}/LookforIDofWikititle_{}'.format(gl_cache, title.replace(u'/', u''))
    if os.path.exists(key):
        with open(key, 'r') as f:
            ret = pickle.load(f)
            return ret

    oldtittle = title

    #title = rectifyRawTitle(title)# preprocess the irregular forms of title

    #lower and compact the title
    lc_title = title
    lc_title = lc_title.replace(u' ',u'')
    lc_title = lc_title.lower()

    if globalDictWikititle2Id.has_key(lc_title) == True:# use LowerCompactTitle to look
        ret = globalDictWikititle2Id.get(lc_title)
    else:#use StandardTitle to look
        firstchar = title[0:1]
        otherchars = title[1:]
        firstchar = firstchar.upper()
        titlefirstcharupper = u'{}{}'.format(firstchar,otherchars)
        if globalDictWikititle2Id.has_key(titlefirstcharupper)==False:
            title_CapitalFirstWord = title.capitalize()
            if globalDictWikititle2Id.has_key(title_CapitalFirstWord)==False:
                #print title_CapitalFirstWord
                title_Titlelize = title.title()
                if globalDictWikititle2Id.has_key((title_Titlelize))==False:
                    title_Titlelize = title_Titlelize.replace(u' Of ', u' of ')
                    title_Titlelize = title_Titlelize.replace(u' The ', u' the ')
                    title_Titlelize = title_Titlelize.replace(u' In ', u' in ')
                    title_Titlelize = title_Titlelize.replace(u' On ', u' on ')
                    title_Titlelize = title_Titlelize.replace(u' To ', u' to ')
                    title_Titlelize = title_Titlelize.replace(u' At ', u' at ')
                    if globalDictWikititle2Id.has_key(title_Titlelize)==False:
                         #print title
                         global globalnonExistingId
                         ret = globalnonExistingId # 99999999 means fail to find its id
                         #                     return ret
                    else:
                        ret = globalDictWikititle2Id[title_Titlelize]
                else:
                    ret = globalDictWikititle2Id[title_Titlelize]
            else:
                ret = globalDictWikititle2Id[title_CapitalFirstWord]
        else:
            ret = globalDictWikititle2Id[titlefirstcharupper]
    with open(key, 'w') as f:
        pickle.dump(ret, f)
    return ret


    #except Exception, reason:
     #   logging.error('{}: There is a serve exception in LookforIDofWikititle(title=\'{}\'), Please check carefully!'.format(get_cur_info(),title))
    #    logging.error("%s:%s" % (reason.__class__.__name__, reason))
      #  print 3/0
       # return None


def ReadMapIdAndLowercompactOrStandardTitle(path_pickle):
    '''
    该函数由ProcWikiParsedOutToTrainCorpus.py拷贝过来.  已有修改，添加了pickle保存变量与读取变量的功能，能有效加快读取速度．
    path_pickle = '/home/ailab/wiki/Datas/WikiExtracter/enwiki-20161001-pages-articles-multistream-index_IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle-pickle.pkl'
    :return:
    '''
    #removedidlist is gotten by function ReadMapWikiIdAndTitle(). if there is a new version Wiki Dump,you can call it to get another removed list
    removedidlist = [49473651,2870157,22545042,670632,1114400,5702430,# enwiki-20161001-pages-articles-multistream-index.txt
                     24350995,13144660,6702435,1189620,2086001,5566745,97100] #Here, I add the known ids , so as to avoid repeat.

    global  globalDictId2StandardWikititle
    global  globalDictWikititle2Id
    globalDictId2StandardWikititle = {} #clear
    globalDictWikititle2Id = {}#clear

    start = default_timer()
    logging.info('Start to prepare globalDictId2StandardWikititle, globalDictCustemId2Title and globalDictWikititle2Id...')

    if os.path.isfile(path_pickle) == True:
        logging.info(u'ReadMapIdAndLowercompactOrStandardTitle from Pickle Dump:\n%s', path_pickle)
        pklfile = file(path_pickle, 'rb')
        #pickle的变量加载load顺序必须要与保存dump顺序完全一致
        globalDictId2StandardWikititle = pickle.load(pklfile)
        globalDictWikititle2Id = pickle.load(pklfile)
        pklfile.close()
        if len(globalDictId2StandardWikititle.keys()) != 16942749 or len(globalDictWikititle2Id.keys()) != 17593298:
            for i in range(1, 20):
                logging.error('{}: "{}" is NOT a RIGHT IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle.'.format(
                        get_cur_info(), path_pickle))
                # return False

    else:
        raise ValueError('cant find path_pickle')

    logging.info('globalDictId2StandardWikititle:{} and globalDictWikititle2Id:{} have been prepared! time:{}s'.format(
        len(globalDictId2StandardWikititle.keys()), len(globalDictWikititle2Id.keys()), default_timer() - start))

    return True




def specialProcessingforErrorwordpair(wold, wnew):
    #from SemEval2017Task2En
    if wold.lower() == u'united nations' and wnew.lower() == u'united nations':
        return u'united nation'
    elif wold.lower() == u'self-driving car' and wnew.lower() == u'self-driving car':
        return u'self-drive car'
    elif wold.lower() == u'snakes and ladders' and wnew.lower() == u'snake and ladders':
        return u'snake and ladder'
    elif wold.lower() == u'robotics' and wnew.lower() == u'robotic':
        return u'robotics'
    elif wold.lower() == u'acoustics' and wnew.lower() == u'acoustic':
        return u'acoustics'
    elif wold.lower() == u'gymnastics' and wnew.lower() == u'gymnastic':
        return u'gymnastics'
    elif wold.lower() == u'thermodynamics' and wnew.lower() == u'thermodynamic':
        return u'thermodynamics'
    elif wold.lower() == u'mercedes' and wnew.lower() == u'mercede':
        return u'mercedes'
    elif wold.lower() == u'rose' and wnew.lower() == u'rise':
        return u'rose'
    elif wold.lower() == u'drawing' and wnew.lower() == u'draw':
        return u'drawing'

    # from WordSim353
    elif wold.lower() == u'memorabilia' and wnew.lower() == u'memorabilium':
        return u'memorabilia'
    #注意353数据集中，media都被还原为了medium。这个是不是不还原也可以？


    return wnew




def ComputeSimiForWordpair_OnlyUseWordForm(w1, w2, flagPreprocFormsOfWordPair, flagMultiwordsvecBySumsinglewordvec, model1, model2 = None):
    '''
    :param w1: the first word
    :param w2: the second word
    :param SimilarityMode: SimilarityMode = Enum('SimilarityMode', ('OnlyUseWordForm','OnlyUseWordConcept'))
    :param model1: the Word2Vec model
    :param model2: the Word2Vec model
    :param flagPreprocFormsOfWordPair: 词对在进行比较之前，是直接使用原样词；还是将它们改为小写；还是将它们进行词形还原并小写
    :param flagMultiwordsvecBySumsinglewordvec: 对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。
    :return: 
    '''
    ret = None
    #assert isinstance(model1, Word2Vec)
    assert isinstance(model1, KeyedVectors)
    if model2 != None:
        #assert isinstance(model2, Word2Vec)
        assert isinstance(model2, KeyedVectors)

    # step1. 如果根据model判断出需要先对wordpair进行词形还原小写处理
    if flagPreprocFormsOfWordPair == PreprocFormsOfWordPair.LemmaLowercase:
        # 取得目标词或短语的 词形还原后的小写形式
        w11 = getLemmasOfText(w1)
        assert len(w11) == 1, u'{}: {} are convert to {}, in  which more than one sentence'.format(get_cur_info(), w1,
                                                                                                   w11)
        w22 = getLemmasOfText(w2)
        assert len(w22) == 1, u'{}: {} are convert to {}, in  which more than one sentence'.format(get_cur_info(), w2,
                                                                                                   w22)
        w11 = u' '.join(w11[0])
        w22 = u' '.join(w22[0])

        w11 = specialProcessingforErrorwordpair(w1, w11)
        w22 = specialProcessingforErrorwordpair(w2, w22)
    elif flagPreprocFormsOfWordPair == PreprocFormsOfWordPair.OnlyLowercase:
        # 只改为小写
        w11 = w1.lower()
        w22 = w2.lower()
    elif flagPreprocFormsOfWordPair == PreprocFormsOfWordPair.OriginalToken:
        # 原样不动
        w11 = w1
        w22 = w2

    if w1 != w11 or w2 != w22:
        # 如果有修改，把它们记录下来，这些才是程序中真正比较的词组，便于后期分析程序
        w1 = w11
        w2 = w22
        logging.debug('actually compared wordpais:    %s    %s', w1, w2)

    # step2. 由model 1 取得短语或单词所对应的向量
    w1vector = getwordvectorForPhraseFromModelOnlybyWordForm(w1, model1, flagMultiwordsvecBySumsinglewordvec)
    canw1vec_bymodel1 = None
    if w1vector is None:  # if w1vector == None:#failed to find
        canw1vec_bymodel1 = False
    else:
        canw1vec_bymodel1 = True
    w2vector = getwordvectorForPhraseFromModelOnlybyWordForm(w2, model1, flagMultiwordsvecBySumsinglewordvec)
    canw2vec_bymodel1 = None
    if w2vector is None:  # if w2vector == None:
        canw2vec_bymodel1 = False
    else:
        canw2vec_bymodel1 = True

    # step3.

    assert canw1vec_bymodel1 != None and canw2vec_bymodel1 != None

    if canw1vec_bymodel1 == True and canw2vec_bymodel1 == True:
        ret = cos_twovectors(w1vector, w2vector)
    elif canw1vec_bymodel1 == True and canw2vec_bymodel1 == False:
        ret = (u' | {}'.format(w2)).encode('utf-8')
    elif canw1vec_bymodel1 == False and canw2vec_bymodel1 == True:
        ret = (u'{} | '.format(w1)).encode('utf-8')
    else:
        ret = (u'{} | {}'.format(w1, w2)).encode('utf-8')

    return ret


def getConceptSetForWordFromdictTitle2ConceptInfoinWiki(word, flagLookupConceptOfWordbyWiki, dictTitle2ConceptInfoinWiki):
    '''
    注意LookupConceptOfWordbyWiki参数，主要提供给getConceptSetForWordFromdictTitle2ConceptInfoinWiki()函数使用。
    这个函数在返回时，并不只是把事先收集的concept.txt中的信息拿过来。它有个简单的排除逻辑：
    只有bNormal为真时，才将Title作为有效概念返回
    只有bRedirect为真且bDisambig为假时，才将RedirectTitle作为有效概念返回
    只有bDisambig为真时，才将DisambigItems作为有效概念返回
    所以Title, RedirectTitle, DisambigItems这三者在任何时刻只可能一个项目返回，不可能出现同时返回的情况。
    当HatnoteLinkItems不为None且长度大于0时，将其作为有效概念返回。如spacecraft的页面开头的ISRO Orbital Vehicle，由HatnoteLinkItems返回。
    当HatnoteLinkDisambigItems不为None且长度大于0时，将其作为有效概念返回。如spacecraft的页面开头的消歧页Spaceship (disambiguation)中的条目，由HatnoteLinkDisambigItems返回的。
    :param word: 
    :param flagLookupConceptOfWordbyWiki: 在查找word所对应的wiki concept时的层次深度(只找Normal的，还是再加上Redirect的，还是再加上歧义页...)
    :param dictTitle2ConceptInfoinWiki: word到wiki concept的转换信息表（由QueryWikiInfoForWordPairFile.py事先准备好的）
    :return: ( {source : list of concepts}, bPageErr )
    '''
    #word = 'dollar'

    #word = 'disk'
    flagLookupConceptOfWordbyWiki = LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
    conceptinfo = None
    assert dictTitle2ConceptInfoinWiki!=None and len(dictTitle2ConceptInfoinWiki)>0
    if dictTitle2ConceptInfoinWiki.has_key(word) == False:
        assert False, u'{}: please check dictTitle2ConceptInfoinWiki, which doesn\'t include: {}!'.format(get_cur_info(), word)
    else:
        conceptinfo = dictTitle2ConceptInfoinWiki[word]

    ret = {}
    if conceptinfo[u'bPageErr'] == True:
        return (None, conceptinfo[u'bPageErr'])
    if conceptinfo[u'bHttpTimeOut'] == True:
        assert False, u'{}: bHttpTimeOut is True for word:{}, please requery its information by Wikipedia'.format(get_cur_info(), word)

    if flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal: #'car'
        # 1)Normal只查找正常页面的Title
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
    elif flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect: #Hadoop
        # 2)Normal_Redirect　只查找正常页面的Title或重定向页的RedirectTitle
        assert not (conceptinfo[u'bNormal']==True and conceptinfo[u'bRedirect'] == True)#因为Normal与Redirect是不相共存的．如果Normal为True，那么Redirect肯定不能为True（应该是None)
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
        elif conceptinfo[u'bRedirect'] == True and conceptinfo[u'bDisambig'] != True:
            #ret.append(conceptinfo[u'RedirectTitle'])
            ret[u'RedirectTitle'] = conceptinfo[u'RedirectTitle']
    elif flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect_Disambig:#disk
        # 3)Normal_Redirect_Disambig #查找正常页面的Title或重定向页的RedirectTitle，并查找歧义页的DisambigItems
        assert not (conceptinfo[u'bNormal']==True and conceptinfo[u'bRedirect'] == True)#因为Normal与Redirect是不相共存的．如果Normal为True，那么Redirect肯定不能为True（应该是None)
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
        elif conceptinfo[u'bRedirect'] == True and conceptinfo[u'bDisambig'] != True:#重定向页　并且　该重定向页不是歧义页,加入重定向的Title
            #ret.append(conceptinfo[u'RedirectTitle'])
            ret[u'RedirectTitle'] = conceptinfo[u'RedirectTitle']

        if conceptinfo[u'bDisambig'] == True:#歧义页(只包含Title或RedirectTitle为歧义页的情况，不含Hatenote中的歧义)
            #ret.extend(conceptinfo[u'DisambigItems'])
            ret[u'DisambigItems'] = conceptinfo[u'DisambigItems']
    elif flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect_HatenoteLinkItems: #gold
        # 4)Normal_Redirect_HatenoteLinkItems　查找正常页面的Title或重定向页的RedirectTitle，并查找页面包含的HatenoteLinkItems
        assert not (conceptinfo[u'bNormal']==True and conceptinfo[u'bRedirect'] == True)#因为Normal与Redirect是不相共存的．如果Normal为True，那么Redirect肯定不能为True（应该是None)
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
        elif conceptinfo[u'bRedirect'] == True and conceptinfo[u'bDisambig'] != True:#重定向页　并且　该重定向页不是歧义页,加入重定向的Title
            #ret.append(conceptinfo[u'RedirectTitle'])
            ret[u'RedirectTitle'] = conceptinfo[u'RedirectTitle']

        if conceptinfo[u'HatnoteLinkItems'] != None and len(conceptinfo[u'HatnoteLinkItems'])!=0:
            #ret.extend(conceptinfo[u'HatnoteLinkItems'])
            ret[u'HatnoteLinkItems'] = conceptinfo[u'HatnoteLinkItems']
    elif flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems:#disk dollar
        # 5)Normal_Redirect_Disambig_HatenoteLinkItems #查找正常页面的Title或重定向页的RedirectTitle，并查找歧义页的DisambigItems
        assert not (conceptinfo[u'bNormal']==True and conceptinfo[u'bRedirect'] == True)#因为Normal与Redirect是不相共存的．如果Normal为True，那么Redirect肯定不能为True（应该是None)
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
        elif conceptinfo[u'bRedirect'] == True and conceptinfo[u'bDisambig'] != True:#重定向页　并且　该重定向页不是歧义页,加入重定向的Title
            #ret.append(conceptinfo[u'RedirectTitle'])
            ret[u'RedirectTitle'] = conceptinfo[u'RedirectTitle']

        if conceptinfo[u'bDisambig'] == True:#歧义页(只包含Title或RedirectTitle为歧义页的情况，不含Hatenote中的歧义)
            #ret.extend(conceptinfo[u'DisambigItems'])
            ret[u'DisambigItems'] = conceptinfo[u'DisambigItems']

        if conceptinfo[u'HatnoteLinkItems'] != None and len(conceptinfo[u'HatnoteLinkItems']) != 0:
            #ret.extend(conceptinfo[u'HatnoteLinkItems'])
            ret[u'HatnoteLinkItems'] = conceptinfo[u'HatnoteLinkItems']
    elif flagLookupConceptOfWordbyWiki == LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems:#dollar, car
        # 6)Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
        # 在Normal_Redirect_Disambig_HatenoteLinkItems的基础上，再增加HatenoteLinkItems所对应的可能的歧义条目．
        # （注意：当bNormal为True时，要把Title也补充到查找结果中去
        #  对比car和disc的wikipedia页面．前者，car是一个normal页，但它的Hatenote中也含有一个消歧页；后者，单纯的一个disambig页．）
        assert not (conceptinfo[u'bNormal']==True and conceptinfo[u'bRedirect'] == True)#因为Normal与Redirect是不相共存的．如果Normal为True，那么Redirect肯定不能为True（应该是None)
        if conceptinfo[u'bNormal'] == True:
            #ret.append(conceptinfo[u'Title'])
            ret[u'Title'] = conceptinfo[u'Title']
        elif conceptinfo[u'bRedirect'] == True and conceptinfo[u'bDisambig'] != True:#重定向页　并且　该重定向页不是歧义页,加入重定向的Title
            #ret.append(conceptinfo[u'RedirectTitle'])
            ret[u'RedirectTitle'] = conceptinfo[u'RedirectTitle']

        if conceptinfo[u'bDisambig'] == True:#歧义页(只包含Title或RedirectTitle为歧义页的情况，不含Hatenote中的歧义)
            #ret.extend(conceptinfo[u'DisambigItems'])
            ret[u'DisambigItems'] = conceptinfo[u'DisambigItems']

        if conceptinfo[u'HatnoteLinkItems'] != None and len(conceptinfo[u'HatnoteLinkItems']) != 0:
            #ret.extend(conceptinfo[u'HatnoteLinkItems'])
            ret[u'HatnoteLinkItems'] = conceptinfo[u'HatnoteLinkItems']
        if conceptinfo[u'HatnoteLinkDisambigItems'] != None and len(conceptinfo[u'HatnoteLinkDisambigItems']) != 0:
            #ret.extend(conceptinfo[u'HatnoteLinkDisambigItems'])
            ret[u'HatnoteLinkDisambigItems'] = conceptinfo[u'HatnoteLinkDisambigItems']
    else:
        assert False, u'{}: need to update source code to handle!'.format(get_cur_info())
    return (ret, conceptinfo[u'bPageErr'])


def findconceptsourceIndictTitle2ConceptInfoinWiki(concept, dictTitle2ConceptInfoinWiki):
    #dictTitle2ConceptInfoinWiki是getConceptSetForWordFromdictTitle2ConceptInfoinWiki()函数的输出。
    #这个函数用来反向查找一个concept的来源是哪里，是title，还是redirect title，还是disambigu.....
    global gl_ConceptSourceFromWikipedia
    #检查一下gl_ConceptSourceFromWikipedia是不是包含了所有的来源，如果不是说明其需要更新一下了
    for sourcename in dictTitle2ConceptInfoinWiki.keys():
        assert sourcename in gl_ConceptSourceFromWikipedia, u'{}: {} is not include in gl_ConceptSourceFromWikipedia, need to update!'.format(get_cur_info(), sourcename)

    retsources= [] #之所以用链表返回，是因为一个concept有可能会出现在多个来源中
    for sourcename in gl_ConceptSourceFromWikipedia:
        if sourcename in dictTitle2ConceptInfoinWiki.keys():
            value = dictTitle2ConceptInfoinWiki[sourcename]
            assert value != None
            if isinstance(value, list):
                if concept in value:
                    retsources.append(sourcename)
            else:
                if concept == value:
                    retsources.append(sourcename)

    if len(retsources)>0:
        return retsources
    else:
        assert False, u'{}: "{}" is not found in dictTitle2ConceptInfoinWiki. There may be a error!'.format(get_cur_info(), concept )


def convertConceptSet2List(ConceptSetForWordFromdictTitle2ConceptInfoinWiki):
    # 将getConceptSetForWordFromdictTitle2ConceptInfoinWiki()函数的字典输出形式，转换为链表
    assert ConceptSetForWordFromdictTitle2ConceptInfoinWiki!=None
    global gl_ConceptSourceFromWikipedia
    for sourcename in ConceptSetForWordFromdictTitle2ConceptInfoinWiki.keys():
        assert sourcename in gl_ConceptSourceFromWikipedia, u'{}: {} is not include in gl_ConceptSourceFromWikipedia, need to update!'.format(get_cur_info(), t)

    ret = []
    for sourcename in gl_ConceptSourceFromWikipedia:
        if sourcename in ConceptSetForWordFromdictTitle2ConceptInfoinWiki.keys():
            value = ConceptSetForWordFromdictTitle2ConceptInfoinWiki[sourcename]
            assert value!=None
            if isinstance(value, list):
                ret.extend(value)
            else:
                ret.append(value)
    #if len(ret)>0:
    #    return ret
    #else:
    #    return None
    return ret

def checkRepeatedElemforList(list):
    #检查链表中存在的重复元素，返回不重复的链表 及 出现重复的元素及次数
    #list = [ 'a', 'b', 'a', 'a', 'c', 'b']
    unrepeatedlist = []
    repeateddict = {}
    assert list!=None and len(list)>0
    for elem in list:
        if elem not in unrepeatedlist:
            unrepeatedlist.append(elem)
        else:
            if repeateddict.has_key(elem):
                repeateddict[elem] += 1
            else:
                repeateddict[elem] = 2
    return (unrepeatedlist, repeateddict)


def normalizeConceptsDict(conceptsDict):
    #将concetpDict中的概念都转换为　词表中的标准概念形式（主要是在前期的程序中，发现Title有时不标准，
    # 比如semeval2017中的multiple sclerosis．第一个m应为大写才是标准形式)
    global gl_ConceptSourceFromWikipedia
    for sourcename in conceptsDict.keys():
        assert sourcename in gl_ConceptSourceFromWikipedia, u'{}: {} is not include in gl_ConceptSourceFromWikipedia, need to update!'.format(get_cur_info(), t)

    new_conceptsDict = {}
    for sourcename in conceptsDict.keys():
        if isinstance(conceptsDict[sourcename], list) == True:
            conceptslist = conceptsDict[sourcename]
            new_conceptslist = []
            for concept in conceptslist:
                id_inWiki = LookforIDofWikititle(concept)# 对Antonín Dvořák查找到其Id: 76572。涉及到读入映射表的过程
                if id_inWiki != globalnonExistingId:
                    ###这个过程与ProWikiParsedOutToTrainCorpus.py中的ParseOneLine2ConceptLemmaOrConcept()是一样的
                    # 对于特殊字符,以bahá'í和brønsted为例(用 人民 and 特殊　也验证了一下．),已验证下列代码。可以将它们从excel中作为词对读出，而后可以在词向量文件中查找到它们对应的词向量
                    # get this id's standard wiki title
                    standardTitle = globalDictId2StandardWikititle.get(id_inWiki)
                    new_conceptslist.append(standardTitle)
                else:
                    str1 = u'{}: "{}" is not found in globalDictWikititle2Id. \n ' \
                          u'this means that online Wikipedia has contained this concept, but your dict file has\'t, may need to update! "index_IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle-pickle.pkl"  '.format(get_cur_info(), concept)
                    logging.info(str1.encode('utf-8'))
                    standardTitle = concept
                    new_conceptslist.append(concept)
                    #assert False, str1
            new_conceptsDict[sourcename] = new_conceptslist
        else:
            assert sourcename == u'Title' or sourcename == u'RedirectTitle'
            concept = conceptsDict[sourcename]
            id_inWiki = LookforIDofWikititle(concept)  # 对Antonín Dvořák查找到其Id: 76572。涉及到读入映射表的过程
            if id_inWiki != globalnonExistingId:
                ###这个过程与ProWikiParsedOutToTrainCorpus.py中的ParseOneLine2ConceptLemmaOrConcept()是一样的
                # 对于特殊字符,以bahá'í和brønsted为例(用 人民 and 特殊　也验证了一下．),已验证下列代码。可以将它们从excel中作为词对读出，而后可以在词向量文件中查找到它们对应的词向量
                # get this id's standard wiki title
                standardTitle = globalDictId2StandardWikititle.get(id_inWiki)
            else:
                str1 = u'{}: "{}" is not found in globalDictWikititle2Id. \n ' \
                      u'this means that oneline Wikipedia has update this concept, but your dict file is still old!"index_IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle-pickle.pkl"  '.format(
                    get_cur_info(), concept)
                #assert False, str
                logging.info(str1)
                standardTitle = concept
            new_conceptsDict[sourcename] = standardTitle

    if new_conceptsDict != conceptsDict:
        logging.debug("conceptsDict: %s", str(conceptsDict))
        logging.debug("new_conceptsDict: %s", str(new_conceptsDict))

    conceptsDict = None
    return new_conceptsDict




def ComputeSimiForWordpair_OnlyUseWordConceptFromWikibyConceptVec(w1,w2,flagMultiwordsvecBySumsinglewordvec,model1,flagLookupConceptOfWordbyWiki, \
                                                                  dictTitle2ConceptInfoinWiki, flagtoSelectFromScoreSetofConceptPair, similarityMode=None):
    '''
    :param w1: the first word
    :param w2: the second word
    :param SimilarityMode: SimilarityMode = Enum('SimilarityMode', ('OnlyUseWordForm','OnlyUseWordConcept'))
    :param model1: the wiki concept2Vec model
    :param flagMultiwordsvecBySumsinglewordvec: 对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。
    :param flagLookupConceptOfWordbyWiki: 在查找word所对应的wiki concept时的层次深度(只找Normal的，还是再加上Redirect的，还是再加上歧义页...)
    :param dictTitle2ConceptInfoinWiki: word到wiki concept的转换信息表（由QueryWikiInfoForWordPairFile.py事先准备好的）
    :param flagtoSelectFromScoreSetofConceptPair: 当两个词的概念集合都找到，并计算完概念对的相似度后，会有很多值，从中选哪个值返回，由这个参数决定
    :return:  返回一个相似度值．如果没有找到概念对　或者　概念对在model中不存在，则返回一些特殊标记字符
    '''
    ret = None
    #assert isinstance(model1, Word2Vec)
    assert isinstance(model1, KeyedVectors)
    assert flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.DontSumAnywords
    assert dictTitle2ConceptInfoinWiki != None
    topK = 20
    # step1.1 根据flagLookupConceptOfWordbyWiki，判断对Word的wiki concept的查找层级,找到对应层级包含的所有概念
    (conceptsDict_w1, bPageErr_w1) = getConceptSetForWordFromdictTitle2ConceptInfoinWiki(w1, flagLookupConceptOfWordbyWiki, dictTitle2ConceptInfoinWiki)
    (conceptsDict_w2, bPageErr_w2) = getConceptSetForWordFromdictTitle2ConceptInfoinWiki(w2, flagLookupConceptOfWordbyWiki, dictTitle2ConceptInfoinWiki)
    #判断词语是否在wikipedia中存在
    if bPageErr_w1 == True and bPageErr_w2 == True:
        logging.info(u'"{}" and "{}" doesn\'t exist in Wikipedia'.format(w1,w2))
        return (u'{} | {}'.format(w1, w2)).encode('utf-8')
    elif bPageErr_w1 == True and bPageErr_w2 != True:
        logging.info(u'"{}" doesn\'t exist in Wikipedia'.format(w1))
        return (u' | {}'.format(w1)).encode('utf-8')
    elif bPageErr_w2 == True and bPageErr_w1 != True:
        logging.info(u'"{}" doesn\'t exist in Wikipedia'.format(w2))
        return (u' | {}'.format(w2)).encode('utf-8')

    assert conceptsDict_w1 != None and conceptsDict_w2 != None
    #将concetpDict中的概念都转换为　词表中的标准概念形式（主要是在前期的程序中，发现Title有时不标准，比如semeval2017中的multiple
    w1_key = u'{}/check_concept_{}'.format(gl_cache, w1.replace(u'/', u''))
    w2_key = u'{}/check_concept_{}'.format(gl_cache, w2.replace(u'/', u''))
    if os.path.exists(w1_key):
        with open(w1_key, 'r') as f:
            conceptsDict_w1 = pickle.load(f)
            conceptsList_w1 = pickle.load(f)
            repeatedconceptList_w1 = pickle.load(f)
    else:
        #sclerosis．第一个m应为大写才是标准形式)
        conceptsDict_w1 = normalizeConceptsDict(conceptsDict_w1)
        # 将dict转换为list，并去重
        conceptsList_w1_tmp = convertConceptSet2List(conceptsDict_w1)
        (conceptsList_w1, repeatedconceptList_w1) = checkRepeatedElemforList(conceptsList_w1_tmp)
        with open(w1_key, 'w') as f:
            pickle.dump(conceptsDict_w1, f)
            pickle.dump(conceptsList_w1, f)
            pickle.dump(repeatedconceptList_w1, f)

    if os.path.exists(w2_key):
        with open(w2_key, 'r') as f:
            conceptsDict_w2 = pickle.load(f)
            conceptsList_w2 = pickle.load(f)
            repeatedconceptList_w2 = pickle.load(f)
    else:
        conceptsDict_w2 = normalizeConceptsDict(conceptsDict_w2)
        conceptsList_w2_tmp = convertConceptSet2List(conceptsDict_w2)
        (conceptsList_w2, repeatedconceptList_w2) = checkRepeatedElemforList(conceptsList_w2_tmp)
        with open(w2_key, 'w') as f:
            pickle.dump(conceptsDict_w2, f)
            pickle.dump(conceptsList_w2, f)
            pickle.dump(repeatedconceptList_w2, f)
    conceptsList_w1_tmp = None; conceptsList_w2_tmp = None
    #判断每个链表中是否都包含有效的concept
    if len(conceptsList_w1) == 0 and len(conceptsList_w2) == 0:
        logging.info(u'"{}" and "{}" can not find valid concept in current condition'.format(w1, w2))
        return (u'{} | {}'.format(w1, w2)).encode('utf-8')
    elif len(conceptsList_w1) == 0 and len(conceptsList_w2) != 0:
        logging.info(u'"{}" can not find valid concept in current condition'.format(w1))
        return (u'{} | '.format(w1)).encode('utf-8')
    elif len(conceptsList_w2) == 0 and len(conceptsList_w1) != 0:
        logging.info(u'"{}" can not find valid concept in current condition'.format(w2))
        return (u' | {}'.format(w2)).encode('utf-8')

    #至此两个词都确定了候选概念的链表
    #str1 = u'dictTitle2ConceptInfoinWiki of "{}": {}'.format(w1, conceptsDict_w1)
    #logging.info(str1.encode('utf-8'))
    str1 = u'conceptsList of "{}"({}): {}'.format(w1, len(conceptsList_w1), listitems2strs(conceptsList_w1, u' | '))
    logging.debug(str1.encode('utf-8'))
    w1_key = u'{}/get_valid_concept_{}'.format(gl_cache, w1.replace(u'/', u''))
    if os.path.exists(w1_key):
        with open(w1_key, 'r') as f:
            validConpectList_w1 = pickle.load(f)
            invalidConceptList_w1 = pickle.load(f)
    else:
        (validConpectList_w1, invalidConceptList_w1) = getValidConceptListinConceptVecModel(conceptsList_w1, model1,
                                                                                            similarityMode)
        with open(w1_key, 'w') as f:
            pickle.dump(validConpectList_w1, f)
            pickle.dump(invalidConceptList_w1, f)
    str1 = u'\t validconceptsList({}): {}'.format(len(validConpectList_w1), listitems2strs(validConpectList_w1, u' | '))
    logging.debug(str1.encode('utf-8'))
    str1 = u'\t invalidconceptsList({}): {}'.format(len(invalidConceptList_w1), listitems2strs(invalidConceptList_w1, u' | '))
    logging.debug(str1.encode('utf-8'))
    #str1 = u'dictTitle2ConceptInfoinWiki of "{}": {}'.format(w2, conceptsDict_w2)
    #logging.info(str1.encode('utf-8'))
    str1 = u'conceptsList of "{}"({}): {}'.format(w2, len(conceptsList_w2), listitems2strs(conceptsList_w2, u' | '))
    logging.debug(str1.encode('utf-8'))
    w2_key = u'{}/get_valid_concept_{}'.format(gl_cache, w2.replace(u'/', u''))
    if os.path.exists(w2_key):
        with open(w2_key, 'r') as f:
            validConpectList_w2 = pickle.load(f)
            invalidConceptList_w2 = pickle.load(f)
    else:
        (validConpectList_w2, invalidConceptList_w2) = getValidConceptListinConceptVecModel(conceptsList_w2, model1,
                                                                                            similarityMode)

        with open(w2_key, 'w') as f:
            pickle.dump(validConpectList_w2, f)
            pickle.dump(invalidConceptList_w2, f)
    str1 = u'\t validconceptsList({}): {}'.format(len(validConpectList_w2), listitems2strs(validConpectList_w2, u' | '))
    logging.debug(str1.encode('utf-8'))
    str1 = u'\t invalidconceptsList({}): {}'.format(len(invalidConceptList_w2), listitems2strs(invalidConceptList_w2, u' | '))
    logging.debug(str1.encode('utf-8'))


    #step1.2 取两个链表的笛卡尔积, 这就是下一步将要依次计算的概念对
    #cartesianProductOfConcepts = [x for x in itertools.product(conceptsList_w1, conceptsList_w2)]
    cartesianProductOfConcepts = [x for x in itertools.product(validConpectList_w1, validConpectList_w2)]
    logging.debug(u'cartesianProduct of two conceptsList includes {} pairs of valid concept({}*{} among total {} pairs {}*{}).'.
                 format(len(cartesianProductOfConcepts), len(validConpectList_w1), len(validConpectList_w2),
                 len(conceptsList_w1)*len(conceptsList_w2), len(conceptsList_w1), len(conceptsList_w2)))


    # step2.
    dictConceptPair2SimScore = {}
    logging.debug(u'begin to compute simi score of each concept pair (only record valid concept in model)...')
    vector_cache = {}
    for (c1, c2) in cartesianProductOfConcepts:
        # step2.1 由model 1 取得每一对可能的concept所对应的概念向量
        if c1 in vector_cache:
            c1vector = vector_cache[c1]
        else:
            c1vector = getconceptvectorForConceptFromModelOnlyUseWordConceptFromWikibyConceptVec(c1, model1, flagMultiwordsvecBySumsinglewordvec, similarityMode)
            vector_cache[c1] = c1vector
        if c2 in vector_cache:
            c2vector = vector_cache[c2]
        else:
            c2vector = getconceptvectorForConceptFromModelOnlyUseWordConceptFromWikibyConceptVec(c2, model1, flagMultiwordsvecBySumsinglewordvec, similarityMode)
            vector_cache[c2] = c2vector

        if c1vector is None:#failed to find
            canc1vec_bymodel1 = False
        else:
            canc1vec_bymodel1 = True
        if c2vector is None:
            canc2vec_bymodel1 = False
        else:
            canc2vec_bymodel1 = True
        # step 2.2 计算向量的余弦相似度，否则 丢弃 或 做ObjectMissing标记
        if canc1vec_bymodel1 == True and canc2vec_bymodel1 == True:
            scorepairc1c2 = cos_twovectors(c1vector, c2vector)
        #elif canc1vec_bymodel1 == True and canc2vec_bymodel1 == False:
        #    scorepairc1c2 = (u' | {}'.format(c2)).encode('utf-8')
        #elif canc1vec_bymodel1 == False and canc2vec_bymodel1 == True:
        #    scorepairc1c2 = (u'{} | '.format(c1)).encode('utf-8')
        #else:
        #    scorepairc1c2 = (u'{} | {}'.format(c1, c2)).encode('utf-8')
        else:
            continue #scorepairc1c2 = gl_simScoreRepresentOneObjectMissing # means that there is a missing concept in (c1,c2) by model1
        # step 2.3 将概念对的 相似度Score存入词典
        dictConceptPair2SimScore[(c1, c2)] = scorepairc1c2
        str1 = u'\t{}    {}    {}'.format(c1, c2, dictConceptPair2SimScore[(c1, c2)])
        logging.debug(str1.encode('utf-8'))

    # step3.
    # step 3.1 sort dictConceptPair2SimScore 对概念对分数词典，按相似度分数降序排列
    sortedListOfPair2Score = sorted(dictConceptPair2SimScore.items(), lambda x, y: cmp(x[1], y[1]), reverse=True)
    # step 3.2 logging the related info  只对有有效分数的概念对做记录
    str1 = []
    countPairWithoutmissingItem = 0
    for item in sortedListOfPair2Score:
        (pair, score) = item
        if score == gl_simScoreRepresentOneObjectMissing:
            continue
        countPairWithoutmissingItem += 1
        (c1, c2) = pair
        strpair = u'{}  {}  {}'.format(c1,c2,score)
        str1.append(strpair)
    if countPairWithoutmissingItem != 0:
        logging.debug(u'Sorted concept pairs On similarity (only both item of pair are found in vector model), includes {} pairs:'.format(len(str1)))
        logging.debug((u'\t' + u'\t|\t'.join(str1)).encode('utf-8'))
        # step 3.3 top-K 个概念对的分数，及概念的来源信息
        logging.debug(u'For top-{} similair concept paris, the source of each item (actually total pairs:{}):'.format(topK, countPairWithoutmissingItem))
        for i in range(0, min(topK, countPairWithoutmissingItem)):#range()函数的返回结果不含右边界
            (pair, score) = sortedListOfPair2Score[i]
            if score == gl_simScoreRepresentOneObjectMissing:
                continue
            (c1, c2) = pair
            c1Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c1, conceptsDict_w1)
            c2Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c2, conceptsDict_w2)
            strpair1 = u'{}  {}  {}'.format(listitems2strs(c1Source,u' | '), listitems2strs(c2Source,u' | '), score)
            strpair2 = u'{}  {}'.format(c1, c2)
            logging.debug(u'\t{}  {}'.format(strpair1, strpair2))
        # step 3.4 从计算出的所有概念对的分数中，选择满足条件的一个分数返回
        (ret, c1, c2, c1Source, c2Source) = SelectSatisfiedOneScoreFromScoreSetOfConceptPairs(sortedListOfPair2Score[0:countPairWithoutmissingItem], conceptsDict_w1, conceptsDict_w2,
                                                          flagtoSelectFromScoreSetofConceptPair)
        if w1!=c1 or w2!=c2:
            str1 = u'{}  {}  {}  ( {}  {} )  ({}  {})'.format(w1, w2, ret, c1, c2, c1Source, c2Source)
        else:
            str1 = u'{}  {}  {}  ({}  {})'.format(w1, w2, ret, c1Source, c2Source)
        logging.debug(str1)
    else:
        logging.debug(u'{}  {}. All of concept pairs in cartesianProduct fail to find concept vector in current Model!'.format(w1, w2))
        ret = None

    if ret == None:
        return (u'missing valid concept').encode('utf-8')


    return ret


def getValidConceptListinConceptVecModel(totalConceptList, ConcetpVecmodel, similarityMode=None):
    #检测totalList中哪些元素 在词向量模型model中存在，哪些不存在，分别返回
    #assert isinstance(ConcetpVecmodel, Word2Vec)
    assert isinstance(ConcetpVecmodel, KeyedVectors)
    assert totalConceptList != None and len(totalConceptList)>0

    validList = []
    invalidList = []

    for item in totalConceptList:
        #借助于这个函数，避免重复编程了。如果concept在model中不存在，这个函数在MultiwordsvecBySumSinglewordvec.DontSumAnywords参数下会返回None。
        conceptvector = getconceptvectorForConceptFromModelOnlyUseWordConceptFromWikibyConceptVec(item, ConcetpVecmodel, MultiwordsvecBySumSinglewordvec.DontSumAnywords, similarityMode)
        #if ConcetpVecmodel.__contains__(item):
        if conceptvector is None:
            invalidList.append(item)
        else:
            validList.append(item)

    return (validList, invalidList)




def SelectSatisfiedOneScoreFromScoreSetOfConceptPairs(sortedListOfPair2Score, conceptsDict_w1, conceptsDict_w2, flagtoSelectFromScoreSetofConceptPair):
    '''
    由ComputeSimiForWordpair_OnlyUseWordConceptFromWikibyConceptVec()函数调用
    
    作用：从候选的sortedListOfPair2Score链表中，选择满足flagtoSelectFromScoreSetofConceptPair条件的一个分数返回。
    当两个词的概念集合都找到，并计算完概念对的相似度后，会有很多值，从中选哪个值返回，由这个函数决定
    返回: 返回 ( maxscore, c1, c2, c1Source, c2Source) 即对于最相似的一对概念，返回其相似度分数、概念1、概念2、概念1来源、概念2来源
    
    sortedListOfPair2Score 这个List的元素形式为(Joule, Cruiser) 0.3957，其一定是已经按照score降序排列的
    
    conceptsDict_w1, conceptsDict_w2 
    两个原始词语的概念信息，由getConceptSetForWordFromdictTitle2ConceptInfoinWiki()函数获取
    注意这两个参数的顺序一定要与sortedListOfPair2Score中的Pair所对应的词的顺序保持一致，不能乱
    
    flagtoSelectFromScoreSetofConceptPair
    1)MAX_ALL 从所有分数中选最大的。不管该concept pair来源是哪里。
    2)MAX_TitleRedirectTitle 限定concept pair的来源为Title或RedirectTitle，从中选择选择分数最高的返回。（这两个来源是可靠性最高的）
    3)MAX_TitleRedirectTitleDisambigItems  限定concept pair来源为Title、RedirectTitle、DisambigItems，从中选择最高的返回。（这三个来源也算可靠性比较高的）
    4)PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_HatnoteLinkItemsHatnoteLinkDisambigItems   如果在Title、RedirectTitle两个来源中能计算出相似度，则优先返回（不论其大还是小）；否则，如果在Title、RedirectTitle、DisambigItems三个来源中能够计算出相似度，则优先返回（不论其大还是小）；否则，扩展到全部范围找。

    '''

    assert sortedListOfPair2Score != None and len(sortedListOfPair2Score) > 0
    assert conceptsDict_w1 != None and conceptsDict_w2 != None
    assert flagtoSelectFromScoreSetofConceptPair != None
    #gl_ConceptSourceFromWikipedia = [u'Title', u'RedirectTitle', u'DisambigItems', u'HatnoteLinkItems', u'HatnoteLinkDisambigItems']
    global gl_ConceptSourceFromWikipedia;
    conceptSourceList = gl_ConceptSourceFromWikipedia #[u'Title', u'RedirectTitle', u'DisambigItems', u'HatnoteLinkItems', u'HatnoteLinkDisambigItems']
    assert conceptSourceList[0]== u'Title' and conceptSourceList[1]==u'RedirectTitle' and conceptSourceList[2]==u'DisambigItems' and \
           conceptSourceList[3]==u'HatnoteLinkItems' and conceptSourceList[4]==u'HatnoteLinkDisambigItems'
    for sourcename in conceptsDict_w1.keys():
        assert sourcename in conceptSourceList, u'{}: {} is not include in conceptSourceList, need to update!'.format(get_cur_info(), sourcename)
    for sourcename in conceptsDict_w2.keys():
        assert sourcename in conceptSourceList, u'{}: {} is not include in conceptSourceList, need to update!'.format(get_cur_info(), sourcename)

    ret = None
    #maxpairscore = None #eg. "(w1, w2) 0.23"
    score = None
    c1 = None
    c2 = None
    c1Source = None
    c2Source = None
    if flagtoSelectFromScoreSetofConceptPair == SelectFromScoreSetofConceptPair.MAX_ALL:
        #1)MAX_ALL 从所有分数中选最大的。不管该concept pair来源是哪里。
        maxpairscore = sortedListOfPair2Score[0] # 取链表0号元素的  概念对 和 分数出来
        (c1, c2) = maxpairscore[0]
        score = maxpairscore[1]
        c1Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c1, conceptsDict_w1)
        c2Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c2, conceptsDict_w2)
        assert c1 != None and c2 != None and score != None and c1Source != None and c2Source != None

    elif flagtoSelectFromScoreSetofConceptPair == SelectFromScoreSetofConceptPair.MAX_TitleRedirectTitle:
        #2)MAX_TitleRedirectTitle 限定concept pair的来源为Title或RedirectTitle，从中选择选择分数最高的返回。（这两个来源是可靠性最高的）
        satisfiedSourcelist = [conceptSourceList[0], conceptSourceList[1]]#u'Title', u'RedirectTitle'
        for pairscore in sortedListOfPair2Score:
            (c1, c2) = pairscore[0] #
            score = pairscore[1]
            #判断其来源是否是满足条件的
            c1Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c1, conceptsDict_w1)
            intersection = set(c1Source).intersection(satisfiedSourcelist)
            if len(intersection) == 0:
                (score, c1, c2, c1Source, c2Source) = (None, None, None, None, None)
                continue
            c2Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c2, conceptsDict_w2)
            intersection = set(c2Source).intersection(satisfiedSourcelist)
            if len(intersection) == 0:
                (score, c1, c2, c1Source, c2Source) = (None, None, None, None, None)
                continue
            #maxpairscore = pairscore
            break #因sortedListOfPair2Score是一个以Score降序排列的链表。由大到小开始找，找到的第一个就是相似度最大的。直接break

    elif flagtoSelectFromScoreSetofConceptPair == SelectFromScoreSetofConceptPair.MAX_TitleRedirectTitleDisambigItems:
        #3)MAX_TitleRedirectTitleDisambigItems 限定concept pair来源为Title、RedirectTitle、DisambigItems，从中选择最高的返回。（这三个来源也算可靠性比较高的）
        satisfiedSourcelist = [conceptSourceList[0], conceptSourceList[1], conceptSourceList[2]]#u'Title', u'RedirectTitle', u'DisambigItems'
        for pairscore in sortedListOfPair2Score:
            (c1, c2) = pairscore[0] #
            score = pairscore[1]
            #判断其来源是否是满足条件的
            c1Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c1, conceptsDict_w1)
            intersection = set(c1Source).intersection(satisfiedSourcelist)
            if len(intersection) == 0:
                (score, c1, c2, c1Source, c2Source) = (None, None, None, None, None)
                continue
            c2Source = findconceptsourceIndictTitle2ConceptInfoinWiki(c2, conceptsDict_w2)
            intersection = set(c2Source).intersection(satisfiedSourcelist)
            if len(intersection) == 0:
                (score, c1, c2, c1Source, c2Source) = (None, None, None, None, None)
                continue
            #maxpairscore = pairscore
            break #因sortedListOfPair2Score是一个以Score降序排列的链表。由大到小开始找，找到的第一个就是相似度最大的。直接break

    elif flagtoSelectFromScoreSetofConceptPair == SelectFromScoreSetofConceptPair.PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_TitleRedirectTitleDisambigItemsHatnoteLinkItemsHatnoteLinkDisambigItems:
        #4)PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_HatnoteLinkItemsHatnoteLinkDisambigItems
        ## 如果在Title、RedirectTitle两个来源中能计算出相似度，则优先返回（不论其大还是小）；
        ## 否则，如果在Title、RedirectTitle、DisambigItems三个来源中能够计算出相似度，则优先返回（不论其大还是小）；否则，扩展到全部范围找。
        logging.debug(u'Under score select condition: {}'.format(flagtoSelectFromScoreSetofConceptPair));
        logging.debug(u'Recursion 1:TRT 2:TRTD 3:ALL ...')
        (score, c1, c2, c1Source, c2Source) = SelectSatisfiedOneScoreFromScoreSetOfConceptPairs(sortedListOfPair2Score, conceptsDict_w1, conceptsDict_w2, \
                                                                  SelectFromScoreSetofConceptPair.MAX_TitleRedirectTitle)
        if score is not None:
            score = score * 1
        else:
            (score, c1, c2, c1Source, c2Source) = SelectSatisfiedOneScoreFromScoreSetOfConceptPairs(sortedListOfPair2Score, conceptsDict_w1, conceptsDict_w2, \
                                                                              SelectFromScoreSetofConceptPair.MAX_TitleRedirectTitleDisambigItems)
            if score is not None:
                score = score * 0.8
            else:
                (score, c1, c2, c1Source, c2Source) = SelectSatisfiedOneScoreFromScoreSetOfConceptPairs(sortedListOfPair2Score, conceptsDict_w1, conceptsDict_w2, \
                                                                      SelectFromScoreSetofConceptPair.MAX_ALL)
                if score is not None:
                    score = score * 0.8


    if score == None:
        logging.info(u'Under score select condition: {}, fails to find valid concept for words!'.format(flagtoSelectFromScoreSetofConceptPair))
        ret = ( None, None, None, None, None )
    else:
        assert c1!=None and c2!=None and score!=None and c1Source!=None and c2Source!=None
        logging.debug(u'Under score select condition: {}'.format(flagtoSelectFromScoreSetofConceptPair))
        #((c1, c2), score) = maxpairscore
        #str1 = u'\tMax score concept pair: {}  {}  {}  {}  {}'.format(maxpairscore[0][0], maxpairscore[0][1], maxpairscore[1], c1Source, c2Source)
        str1 = u'\tMax score concept pair: {}  {}  {}  {}  {}'.format(c1, c2, score, c1Source, c2Source)
        logging.debug(str1.encode('utf-8'))
        #ret = maxpairscore[1]
        ret = ( score, c1, c2, c1Source, c2Source )

    return ret


def ComputeSimiForWordpair(w1, w2, similarityMode, flagPreprocFormsOfWordPair, flagMultiwordsvecBySumsinglewordvec, model1, model2 = None, \
                           flagLookupConceptOfWordbyWiki = None, dictTitle2ConceptInfoinWiki = None, flagtoSelectFromScoreSetofConceptPair = None):
    '''
    :param w1: the first word
    :param w2: the second word
    :param SimilarityMode: SimilarityMode = Enum('SimilarityMode', ('OnlyUseWordForm','OnlyUseWordConcept'))
    :param model1: the Word2Vec model
    :param model2: the Word2Vec model
    :param flagPreprocFormsOfWordPair: 词对在进行比较之前，是直接使用原样词；还是将它们改为小写；还是将它们进行词形还原并小写
    :param flagMultiwordsvecBySumsinglewordvec: 对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。
    :param flagLookupConceptOfWordbyWiki: 在查找word所对应的wiki concept时的层次深度(只找Normal的，还是再加上Redirect的，还是再加上歧义页...)
    :param dictTitle2ConceptInfoinWiki: word到wiki concept的转换信息表（由QueryWikiInfoForWordPairFile.py事先准备好的）
    :param flagtoSelectFromScoreSetofConceptPair: 当两个词的概念集合都找到，并计算完概念对的相似度后，会有很多值，从中选哪个值返回，由这个参数决定 
    :return: 
    '''
    ret = None

    if similarityMode == SimilarityMode.OnlyUseWordForm:
        ret = ComputeSimiForWordpair_OnlyUseWordForm(w1,w2,flagPreprocFormsOfWordPair,flagMultiwordsvecBySumsinglewordvec,model1,model2)
    elif similarityMode == SimilarityMode.OnlyUseWordConceptFromWikibyConceptVec or similarityMode == SimilarityMode.OnlyUseWordConceptFromNode2vec:
        ret = ComputeSimiForWordpair_OnlyUseWordConceptFromWikibyConceptVec(w1,w2,flagMultiwordsvecBySumsinglewordvec,model1,\
                                                                            flagLookupConceptOfWordbyWiki, dictTitle2ConceptInfoinWiki, \
                                                                            flagtoSelectFromScoreSetofConceptPair, similarityMode)
    elif similarityMode == SimilarityMode.UseWordFormAndWordConcept or similarityMode == SimilarityMode.OnlyUserCombineWikiConceptNode2vec:
        retWordForm = ComputeSimiForWordpair_OnlyUseWordForm(w1, w2, flagPreprocFormsOfWordPair,
                                                     flagMultiwordsvecBySumsinglewordvec, model1, model2)
        retWordConcept = ComputeSimiForWordpair_OnlyUseWordConceptFromWikibyConceptVec(w1, w2, flagMultiwordsvecBySumsinglewordvec,model1, \
                                                                            flagLookupConceptOfWordbyWiki,dictTitle2ConceptInfoinWiki, \
                                                                            flagtoSelectFromScoreSetofConceptPair)

        if isinstance(retWordConcept, float) or isinstance(retWordConcept, int):
            return retWordConcept
        elif isinstance(retWordForm, float) or isinstance(retWordForm, int):
            return retWordForm
        else:
            return 'WordForm: {}, WordConcept: {}'.format(retWordForm, retWordConcept)
    else:
        assert False, u'{}: need to update source code to handle {}'.format(get_cur_info(), similarityMode)


    return ret


def getflagPreprocFormsOfWordPair_FromFromEmbeddingfilePath(embeddingfilepath):
    # 推测词对在进行比较之前，是直接使用原样词；还是将它们改为小写；还是将它们进行词形还原并小写
    # 0421121543-OL-OLOS-SI50_al0.025_wi5_mc5_mvnone_sa0.001_se1_WO11_ma0.0001_sg0_hs0_ne5_cb1_hahash_it5_nw0_trnone_sv1_bw10000-!OriginalLemma.Merged.txt.gen.syn1neg
    # OriginalToken, OriginalLemma, ConceptToken, ConceptLemma     ProcWikiParsedOutToTrainCorpus.py
    # OT, OL, CT, CL
    # 其中OT，CT，都是Token，都是区分大小写的
    # 其中OL，CL，都是还原后的，已不再区分大小写
    global PreprocFormsOfWordPair
    filename = os.path.basename(embeddingfilepath)

    part = filename[0:20]
    if part.find(u'-OL-') != -1:
        return PreprocFormsOfWordPair.LemmaLowercase
    elif part.find(u'-CL-') != -1:
        return PreprocFormsOfWordPair.LemmaLowercase
    elif part.find(u'-OT-') != -1:
        return PreprocFormsOfWordPair.OriginalToken
    elif part.find(u'-CT-') != -1:
        return PreprocFormsOfWordPair.OriginalToken
    else:
        assert False, u'{}: need to update source code'.format_map(get_cur_info())



def getbLowerconvertFromEmbeddingfilePath(embeddingfilepath):
    #0421121543-OL-OLOS-SI50_al0.025_wi5_mc5_mvnone_sa0.001_se1_WO11_ma0.0001_sg0_hs0_ne5_cb1_hahash_it5_nw0_trnone_sv1_bw10000-!OriginalLemma.Merged.txt.gen.syn1neg
    #OriginalToken, OriginalLemma, ConceptToken, ConceptLemma     ProcWikiParsedOutToTrainCorpus.py
    #OT, OL, CT, CL
    #其中OT，CT，都是Token，都是区分大小写的
    #其中OL，CL，都是还原后的，已不再区分大小写
    filename = os.path.basename(embeddingfilepath)

    part = filename[0:20]
    if part.find(u'-OL-')!=-1:
        return True
    elif part.find(u'-CL-')!=-1:
        return True
    elif part.find(u'-OT-')!=-1:
        return False
    elif part.find(u'-CT-')!=-1:
        return False
    else:
        assert False, u'{}: need to update source code'.format_map(get_cur_info())


def isNumber(value):
    '''判断变量value是否为数字（浮点、整型...都可以）'''
    try:
        value + 1
    except TypeError:
        return False
    else:
        return True


def ComputeSimilarityForOneDataSet(dataSet, similarityMode, embeddingFile1, embeddingFile2, flagPreprocFormsOfWordPair, \
                                   flagMultiwordsvecBySumsinglewordvec, recTime, flagLookupConceptOfWordbyWiki, flagtoSelectFromScoreSetofConceptPair):
    '''
    
    :param dataSet:  WordSim353 , SemEval2017Task2En 
    :param similarityMode: OnlyUseWordForm, ...
    :param embeddingFile1: the path of embedding file1 使用哪个词向量文件，根据词向量文件的结尾字符.vec, .gen, .bin来决定使用哪种词向量读入函数。(.vec代表google word2vec的text模式，可直接看到向量值；.bin代表google word2vec的binary模式；.gen代表gensim自己的模式)
    :param embeddingFile2: 
    :param flagPreprocFormsOfWordPair: 对在进行比较之前，是直接使用原样词；还是将它们改为小写；还是将它们进行词形还原并小写
    :param flagMultiwordsvecBySumsinglewordvecrecTime: 对于多字短语的词向量如何求，是否允许直接累加其包含的单个词的词向量作为结果。
    :param recTime:  the time str that to be recorded in excel file
    :param flagLookupConceptOfWordbyWiki: 在查找word所对应的wiki concept时的层次深度(只找Normal的，还是再加上Redirect的，还是再加上歧义页...)
    :param flagtoSelectFromScoreSetofConceptPair: 当两个词的概念集合都找到，并计算完概念对的相似度后，会有很多值，从中选哪个值返回，由这个参数决定 
    :return: 
    '''
    global gl_wordsim353path, gl_wordsemeval2017task2enpath, gl_rg65path, gl_men3000path, gl_mturk771path, gl_rw2034path, \
        gl_ws353simpath, gl_ws353relpath, gl_mc30path, gl_modelembedding1, gl_modelembedding2
    global gl_conceptws353relpath, gl_conceptws353simpath, gl_conceptrw2034path, gl_conceptmturk771path, gl_conceptmen3000path, \
        gl_conceptrg65path, gl_conceptsemeval2017task2enpath, gl_conceptwordsim353path, gl_conceptmc30path
    #global gl_flagLookupConceptOfWordbyWiki

    # 记录开始处理的时间
    extract_start = default_timer()
    logging.info("Begin to Compute Similairty for DataSet: %s", dataSet)

    #get path of data file
    #excelScorePath 数据库Excel表的路径
    #wikiConceptPath 提前由wikipedia查好的word对应的wiki concept文件路径
    (excelScorePath, wikiConceptPath) = GetAndCheckDataFile(dataSet)

    # Step1. read the word pairs from Excel file
    wb = load_workbook(excelScorePath)
    ws = wb.get_sheet_by_name(u'RunRecords')

    dictTitle2ConceptInfoinWiki = None
    #if need, read the map info from Title to Wiki Concepts
    if similarityMode == SimilarityMode.OnlyUseWordConceptFromWikibyConceptVec or similarityMode == SimilarityMode.UseWordFormAndWordConcept\
            or similarityMode == SimilarityMode.OnlyUseWordConceptFromNode2vec or similarityMode == SimilarityMode.OnlyUserCombineWikiConceptNode2vec:
        dictTitle2ConceptInfoinWiki = readLookupFileWord2ConceptbyWiki(wikiConceptPath)

    # Step1.1 check the format of data file
    # CheckFormatandDatainExcelfile(ws, dataSet)

    # the new column that is ready to save the new data
    newcolumnid = ws.max_column + 1

    # Step 1.2 read data from Excel file
    # the map between word pair and its rowid, so as to which row to write the output
    dict_string2rowid = {}
    # the map between word pair and its Human score
    dict_wordpair2humanscore = {}
    # the map between word pair and autoRun score
    dict_wordpair2autorunscore = {}
    # the order of word pair
    list_wordpairs = []
    # read the data from Excel file, and save them in the above map
    for rowid in range(1, ws.max_row + 1):
        a = ws.cell('%s%s' % ('A', rowid)).value
        b = ws.cell('%s%s' % ('B', rowid)).value
        c = ws.cell('%s%s' % ('C', rowid)).value
        if a == u'Word 1' and b == u'Word 2' and c == u'Human (mean)':
            continue
        if a == None and b == None and c != None:
            dict_string2rowid[c] = rowid
        elif a != None and b != None and c != None:
            key = u'{}\t{}'.format(a, b)
            if key in dict_string2rowid.keys():
                logging.info(u'{}: there is a repeat key "{}" in dataset'.format(get_cur_info(), key))
                if dataSet == DataSet.WordSim353 and key == u'money\tcash':
                    pass
                else:
                    assert False, u'{}: there is a unknown repeat key "{}" in dataset!'.format(get_cur_info(),key)
            dict_string2rowid[key] = rowid
            dict_wordpair2humanscore[key] = c
            list_wordpairs.append(key)
        else:
            assert False, u'{}: need to update the source code to handle this'.format(get_cur_info())


    # Step 2. Get the similarity  score of autoRun
    modelloadbegintime = default_timer()

    if gl_modelembedding1 == None or (gl_modelembedding2 == None and embeddingFile2 != None):
        logging.info(u'begin to load model1... \n%s\n%.1fMb', os.path.basename(embeddingFile1),
                     os.path.getsize(embeddingFile1) / 1024.0 / 1024)
        gl_modelembedding1 = loadModelforWord2Vec(embeddingFile1)
        if (gl_modelembedding2 == None and embeddingFile2 != None):
            logging.info(u'begin to load model2... \n%s\n%.1fMb', os.path.basename(embeddingFile2),
                         os.path.getsize(embeddingFile2) / 1024.0 / 1024)
            gl_modelembedding2 = u''
            assert False, u'{}: need to update source code'.format_map(get_cur_info())
        logging.info('finish to load model in %.1f seconds', default_timer() - modelloadbegintime)

    #for each pair to compute its similarity
    count = 0
    for t in list_wordpairs:
        count += 1
        # t = u'money\tcash'
        # t = u'El Niño\tequator'
        wordpair = t
        words = wordpair.split(u'\t')
        assert len(words) == 2, u'{}: the wordpair "{}" seems a error, please check it!'.format(get_cur_info(), wordpair)
        logging.info("%d/%d    %s    %s    ...", count, len(list_wordpairs), \
                     words[0].encode('utf-8'), words[1].encode('utf-8'))

        #score = ComputeSimiForWordpair(u'中国', u'特殊', similarityMode, modelembedding1, modelembedding2)
        #score = ComputeSimiForWordpair(u'特殊特殊特殊', u'特殊', similarityMode, modelembedding1, modelembedding2)
        with Timer() as timer:
            # words = [u'boy', u'lad']
            score = ComputeSimiForWordpair(words[0], words[1], similarityMode, flagPreprocFormsOfWordPair, flagMultiwordsvecBySumsinglewordvec, gl_modelembedding1, gl_modelembedding2, flagLookupConceptOfWordbyWiki, dictTitle2ConceptInfoinWiki, flagtoSelectFromScoreSetofConceptPair)
        logging.info("%d/%d    %s    %s    cost time: %fs", count, len(list_wordpairs), words[0].encode('utf-8'), words[1].encode('utf-8'), timer.cost)
        #### this code block is to compute the similarity score with self-defined cos_twovectors() function. whose result is same with word2vec.model.similarity()
        #if isNumber(score):
        #    print score
        #    em1 = gl_modelembedding1[words[0]]
        #    em2 = gl_modelembedding1[words[1]]
        #    print cos_twovectors(em1,em2)
        #    print cos_twovectors(em1, em2+em2)
        ####



        logging.info("%d/%d    %s    %s    %s    %s\n\n", count, len(list_wordpairs), \
                     words[0].encode('utf-8'), words[1].encode('utf-8'), dict_wordpair2humanscore[wordpair], score)

        assert wordpair in dict_wordpair2humanscore.keys(), u'{}: {}'.format(get_cur_info(), wordpair)
        dict_wordpair2autorunscore[wordpair] = score
    dataSet_count = count


    # Step 3. Save the information of autoRun, such as parameter, scores
    # save the score in Excel
    recTime = recTime.encode('utf-8')
    recSimilarityMode = (u'{}'.format(similarityMode)[u'{}'.format(similarityMode).find(u'.')+1:]).encode('utf-8')
    recEmbeddingFile1 = os.path.basename(embeddingFile1).encode('utf-8')
    if embeddingFile2 != None:
        recEmbeddingFile2 = os.path.basename(embeddingFile2).encode('utf-8')
    else:
        recEmbeddingFile2 = u'None'.encode('utf-8')
    recflagPreprocFormsOfWordPair = (u'{}'.format(flagPreprocFormsOfWordPair)[u'{}'.format(flagPreprocFormsOfWordPair).find(u'.')+1:]).encode('utf-8')
    recflagMultiwordsvecBySumsinglewordvec = (u'{}'.format(flagMultiwordsvecBySumsinglewordvec)[u'{}'.format(flagMultiwordsvecBySumsinglewordvec).find(u'.')+1:]).encode('utf-8')
    recflagLookupConceptOfWordbyWiki = (u'{}'.format(flagLookupConceptOfWordbyWiki)[u'{}'.format(flagLookupConceptOfWordbyWiki).find(u'.')+1:]).encode('utf-8')
    recflagtoSelectFromScoreSetofConceptPair = (u'{}'.format(flagtoSelectFromScoreSetofConceptPair)[u'{}'.format(flagtoSelectFromScoreSetofConceptPair).find(u'.')+1:]).encode('utf-8')
    #print list_wordpairs
    humanscores = getHumanScoreOrderedbyWordpairs(dict_wordpair2humanscore, list_wordpairs, dataSet)
    #print humanscores
    autorunscores = getAutorunScoreOrderedbyWordpairs(dict_wordpair2autorunscore, list_wordpairs, dataSet)
    #print autorunscores
    [spearman_r, spearman_pvalue, spearman_note, spearman_numofvalidpairs] = computeSpearmanr(copy.deepcopy(humanscores), copy.deepcopy(autorunscores), \
                                                                    bRemoveNonNumberinBothlist=True, bFallbackForNonNumberwithMiddlepointscore=False)
    [pearson_r, pearson_pvalue, pearson_note, pearson_numofvalidpairs] = computePearsonr(copy.deepcopy(humanscores), copy.deepcopy(autorunscores), \
                                                                bRemoveNonNumberinBothlist=True, bFallbackForNonNumberwithMiddlepointscore=False)
    recSpearmanCorWithHuman = spearman_r
    recSpearman_pvalue = spearman_pvalue
    recPearsonCorWithHuman = pearson_r
    recPearson_pvalue = pearson_pvalue
    recNotes = None
    if spearman_note != None:
        recNotes = spearman_note
    if pearson_note != None:
        recNotes = recNotes + u'  ' + pearson_note
    if recNotes == None:
        recNotes = u'None'
    recNotes = recNotes.encode('utf-8')

    [spearman_r_rplNN05, spearman_pvalue_rplNN05, spearman_note_rplNN05, _] = computeSpearmanr(copy.deepcopy(humanscores), copy.deepcopy(autorunscores),\
                                                                                            bRemoveNonNumberinBothlist=False, bFallbackForNonNumberwithMiddlepointscore=True,\
                                                                                            middelpointscoreinlist=0.5)
    [pearson_r_rplNN05, pearson_pvalue_rplNN05, pearson_note_rplNN05, _] = computePearsonr(copy.deepcopy(humanscores), copy.deepcopy(autorunscores), \
                                                                                        bRemoveNonNumberinBothlist=False, bFallbackForNonNumberwithMiddlepointscore=True, \
                                                                                        middelpointscoreinlist=0.5)
    recSpearmanCorWithHuman_rplNN05 = spearman_r_rplNN05
    recSpearman_pvalue_rplNN05 = spearman_pvalue_rplNN05
    recPearsonCorWithHuman_rplNN05 = pearson_r_rplNN05
    recPearson_pvalue_rplNN05 = pearson_pvalue_rplNN05
    recNotes_rplNN05 = None
    if spearman_note_rplNN05 != None:
        recNotes_rplNN05 = spearman_note_rplNN05
    if pearson_note_rplNN05 != None:
        recNotes_rplNN05 = recNotes_rplNN05 + u'  ' + pearson_note_rplNN05
    if recNotes_rplNN05 == None:
        recNotes_rplNN05 = u'None'
    recNotes_rplNN05 = recNotes_rplNN05.encode('utf-8')


    # 3.1 save parameters
    #col_letter = get_column_letter(newcolumnid)
    col_letter = int_toletter_columnExcel(newcolumnid)
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Time'))
    ws.cell(colrow).value = recTime
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'SimilarityMode'))
    ws.cell(colrow).value = recSimilarityMode
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'EmbeddingFile1'))
    ws.cell(colrow).value = recEmbeddingFile1
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'EmbeddingFile2'))
    ws.cell(colrow).value = recEmbeddingFile2
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'flagPreprocFormsOfWordPair'))
    ws.cell(colrow).value = recflagPreprocFormsOfWordPair #u'True'.encode('utf-8') if bflagconvert2LemmaLowercaseforwordpair == True else u'False'.encode('utf-8')
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'flagMultiwordsvecBySumsinglewordvec'))
    ws.cell(colrow).value = recflagMultiwordsvecBySumsinglewordvec
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'flagLookupConceptOfWordbyWiki'))
    ws.cell(colrow).value = recflagLookupConceptOfWordbyWiki
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'flagtoSelectFromScoreSetofConceptPair'))
    ws.cell(colrow).value = recflagtoSelectFromScoreSetofConceptPair
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'SpearmanCorWithHuman'))
    ws.cell(colrow).value = recSpearmanCorWithHuman
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Spearman-pvalue'))
    ws.cell(colrow).value = recSpearman_pvalue
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'PearsonCorWithHuman'))
    ws.cell(colrow).value = recPearsonCorWithHuman
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Pearson-pvalue'))
    ws.cell(colrow).value = recPearson_pvalue
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Notes'))
    ws.cell(colrow).value = recNotes
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'SpearmanCorWithHuman_ReplaceNonNumber0.5'))
    ws.cell(colrow).value = recSpearmanCorWithHuman_rplNN05
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Spearman-pvalue_ReplaceNonNumber0.5'))
    ws.cell(colrow).value = recSpearman_pvalue_rplNN05
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'PearsonCorWithHuman_ReplaceNonNumber0.5'))
    ws.cell(colrow).value = recPearsonCorWithHuman_rplNN05
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Pearson-pvalue_ReplaceNonNumber0.5'))
    ws.cell(colrow).value = recPearson_pvalue_rplNN05
    colrow = '%s%s' % (col_letter, dict_string2rowid.get(u'Notes_ReplaceNonNumber0.5'))
    ws.cell(colrow).value = recNotes_rplNN05
    #wb.save(excelScorePath)
    # 3.2 save the similarity score of each pair
    #col_letter = get_column_letter(newcolumnid)
    for t in list_wordpairs:
        rowid = dict_string2rowid.get(t)
        colrow = '%s%s' % (col_letter, rowid)
        ws.cell(colrow).value = dict_wordpair2autorunscore.get(t)

    # if dataSet == DataSet.WordSim353: # there is a repeated "money\tcash"
    #     #when need to update, the three '49' need to change at same time
    #     rowspecial = 51
    #     assert ws.cell("A"+str(rowspecial)).value == u'money' and ws.cell(
    #         "B"+str(rowspecial)).value == u'cash', u'{}: Please assure the content of wordsim353score.xlsx is right!'.format(
    #         get_cur_info())
    #     ws.cell('%s%s' % (col_letter, rowspecial)).value = dict_wordpair2autorunscore.get('money\tcash')

    # 3.3 save Excel file
    wb.save(excelScorePath)

    # 3.4 save its result to AllscoreCompare.xlsx
    #SaveCorScores2compareexcel(dataSet, SpearmanCorWithHuman, PearsonCorWithHuman, Notes, Time, SimilarityMode, EmbeddingFile1, EmbeddingFile2, flagPreprocFormsOfWordPair, flagMultiwordsvecBySumsinglewordvec)
    SaveCorScores2compareexcel(dataSet, recSpearmanCorWithHuman_rplNN05, recPearsonCorWithHuman_rplNN05,  recNotes_rplNN05,
                               recTime, recSimilarityMode, recEmbeddingFile1, recEmbeddingFile2,
                               recflagPreprocFormsOfWordPair, recflagMultiwordsvecBySumsinglewordvec,
                               recflagLookupConceptOfWordbyWiki, recflagtoSelectFromScoreSetofConceptPair)
    # 计算运行时间
    extract_duration = default_timer() - extract_start
    #logging.info("The autorun's output\nWord 1\tWord 2\tAutoScore")
    #logging.info('\n'+getStrOfdictWordpair2Score(dict_wordpair2autorunscore, list_orderWordpair=list_wordpairs))
    logging.info("Congratulations!")
    logging.info("Finished computing similarity of %d word pairs in %s dataset in %.1f seconds",
                 len(list_wordpairs), dataSet, extract_duration)
    logging.info("Spearman:%.4f Pearson:%.4f , actually valid number of word pair:%d", spearman_r, pearson_r, \
                 countNumberElementsinList(autorunscores)) # autorunscores has been removed in [pearson_r, pearson_pvalue, pearson_note] = computePearsonr(humanscores, autorunscores, bRemoveNonNumberinBothlist=True)
    logging.info("If replace the nonNumber in scorelist with middelpiont, Spearman:%.4f Pearson:%.4f , actually replaced number of word pairs:%d", spearman_r_rplNN05, pearson_r_rplNN05, \
                 len(autorunscores)-countNumberElementsinList(autorunscores))

    assert spearman_numofvalidpairs == pearson_numofvalidpairs
    #返回[数据集名字, 实际有效词对的spearman系数, 实际有效词对的pearson系数, 参与相关计算的有效词对数量, 数据集包含的词对数量]
    return (dataSet, spearman_r, pearson_r, spearman_numofvalidpairs, dataSet_count,
            recTime, recSimilarityMode, recEmbeddingFile1, recEmbeddingFile2, recflagPreprocFormsOfWordPair,
             recflagMultiwordsvecBySumsinglewordvec, recflagLookupConceptOfWordbyWiki, recflagtoSelectFromScoreSetofConceptPair)


def cos_twovectors(vector1, vector2):
    '''
    compute cosine similarity of two vectors
    :param vector1: 
    :param vector2: 
    :return: 
    '''
    return numpy.dot(matutils.unitvec(vector1), matutils.unitvec(vector2))


class Wordpair(object):
    def __init__(self, embedding_file, pkl_file):
        self.embedding_file = embedding_file
        self.pkl_file = pkl_file
        self.concept_file = './data/concept.txt'
        self.is_init = False
        self.similarityMode = SimilarityMode.UseWordFormAndWordConcept
        self.flagPreprocFormsOfWordPair = PreprocFormsOfWordPair.LemmaLowercase
        self.flagMultiwordsvecBySumsinglewordvec = MultiwordsvecBySumSinglewordvec.DontSumAnywords
        self.flagLookupConceptOfWordbyWiki = LookupConceptOfWordbyWiki.Normal_Redirect_Disambig_HatenoteLinkItems_HatenoteLinkDisambigItems
        self.flagtoSelectFromScoreSetofConceptPair = SelectFromScoreSetofConceptPair.PRIORITY_TitleRedirectTitle_TitleRedirectTitleDisambigItems_TitleRedirectTitleDisambigItemsHatnoteLinkItemsHatnoteLinkDisambigItems
        self.dictTitle2ConceptInfoinWiki = {}

    def init(self):
        if self.is_init:
            return True
        # 初始化操作
        print 'load pkl'
        ReadMapIdAndLowercompactOrStandardTitle(self.pkl_file)
        print 'load concept'
        self.dictTitle2ConceptInfoinWiki = readLookupFileWord2ConceptbyWiki(self.concept_file)
        print 'load embedding'
        self.embedding = loadModelforWord2Vec(self.embedding_file)
        self.is_init = True

    def score(self, word1, word2):
        self.init()
        score = ComputeSimiForWordpair(word1, word2, self.similarityMode, self.flagPreprocFormsOfWordPair,
                                       self.flagMultiwordsvecBySumsinglewordvec, self.embedding, None,
                                       self.flagLookupConceptOfWordbyWiki, self.dictTitle2ConceptInfoinWiki,
                                       self.flagtoSelectFromScoreSetofConceptPair)
        return score

def main():


    parser = argparse.ArgumentParser(prog=os.path.basename(sys.argv[0]),
                                     formatter_class=argparse.RawDescriptionHelpFormatter,
                                     description=__doc__)
    parser.add_argument("DataSet", #All RG65　MEN3000　MTURK771　RW2034　WS353SIM　WS353REL
                        help="to assign the dataset to compare similairty (WordSim353, SemEval2017Task2En")
    parser.add_argument("SimilarityMODE", #OnlyUseWordForm OnlyUseWordConceptFromWikibyConceptVec UseWordFormAndWordConcept
                        help="the mode to compute similarity with WordForm or WordConcept (OnlyUseWordForm, OnlyUseWordConceptFromWikibyConceptVec, UseWordFormAndWordConcept)")
    parser.add_argument("EmbeddingFile1",
                        help="the Word Embedding file from Wikipedia")
    #其它参数gl_flagPreprocFormsOfWordPair、gl_flagMultiwordsvecBySumsinglewordvec在程序中直接改参数

    groupS = parser.add_argument_group('Special')
    groupS.add_argument("-e","--EmbeddingFile2",default="None",
                        help="the Word Embedding file from Wikipedia")
    groupS.add_argument("-q", "--quiet", action="store_true",
                        help="suppress reporting progress info")
    groupS.add_argument("--debug", action="store_true",
                        help="print debug info")
    groupS.add_argument("--cache", default="cache",
                        help="use cache")

    args = parser.parse_args()


    ###
    FORMAT = '%(levelname)s %(asctime)s: %(message)s'
    logging.basicConfig(format=FORMAT)
    logger = logging.getLogger()#create a logger
    if not args.quiet:
        logger.setLevel(logging.INFO)  #DEBUG, INFO, WARNING, ERROR, CRITICAL
    if args.debug:
        logger.setLevel(logging.DEBUG)
    handlerfile = logging.FileHandler(os.path.basename(__file__) + '.log')  # create a handler, to write to log file
    #handlercon = logging.StreamHandler()#create a handler, to write to console
    handlerfile.setLevel(logging.DEBUG)
    #handlercon.setLevel(logging.INFO)
    handlerformat = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')#define the formation of handler
    handlerfile.setFormatter(handlerformat)
    #handlercon.setFormatter(handlerformat)
    logger.addHandler(handlerfile)#add handler to logger
    #logger.addHandler(handlercon)
    ###

    # job_quere = Queue()
    # out_quere = Queue()
    # workers_num = cpu_count() - 1
    # workers = []
    # for i in range(workers_num):
    #     computer = Process(target=ComputeProcess,
    #                        args=(job_quere, out_quere))
    #     computer.daemon = True
    #     computer.start()
    #     workers.append(computer)



    global SimilarityMode, DataSet, gl_wordsim353path, gl_wordsemeval2017task2enpath, gl_bMultiwordsvecBySumsinglewordvec, gl_flagLookupConceptOfWordbyWiki,gl_cache
    list_dataSet = []
    list_dataSetexcelfile = []

    gl_cache = args.cache
    HandleInputParameterDataSet(args.DataSet, list_dataSet, list_dataSetexcelfile)

    #similarityMode = None
    if args.SimilarityMODE == u'OnlyUseWordForm':
        similarityMode = SimilarityMode.OnlyUseWordForm
    elif args.SimilarityMODE == u'OnlyUseWordConceptFromWikibyConceptVec':
        similarityMode = SimilarityMode.OnlyUseWordConceptFromWikibyConceptVec
    elif args.SimilarityMODE == u'UseWordFormAndWordConcept':
        similarityMode = SimilarityMode.UseWordFormAndWordConcept
    elif args.SimilarityMODE == u'OnlyUseWordConceptFromNode2vec':
        similarityMode = SimilarityMode.OnlyUseWordConceptFromNode2vec
    elif args.SimilarityMODE == u'OnlyUserCombineWikiConceptNode2vec':
        similarityMode = SimilarityMode.OnlyUserCombineWikiConceptNode2vec
    else:
        assert False, u'{}: need to update handle code! input parameter: {}'.format(get_cur_info(), args.SimilarityMode)


    embeddingFile1 = args.EmbeddingFile1
    assert os.path.isfile(embeddingFile1) == True, u'{}: {} is not a real Embedding file'.format(get_cur_info(), embeddingFile1)
    embeddingFile2 = args.EmbeddingFile2
    if embeddingFile2!=u'None':
        assert os.path.isfile(embeddingFile2) == True, u'{}: {} is not a real Embedding file'.format(get_cur_info(), embeddingFile2)
    else:
        embeddingFile2 = None
    advicflagPreprocFormsOfWordPair = getflagPreprocFormsOfWordPair_FromFromEmbeddingfilePath(embeddingFile1)
    #bflagconvert2LemmaLowercaseforwordpair = getbLowerconvertFromEmbeddingfilePath(embeddingFile1)



    recTime = get_time_info()#the time to record

    logging.info(u'{}'.format('\n' * 10 + '=' * 40 + '\nBegin to compute similarity of word pair ...'))
    logging.info(u'cache path: {}'.format(gl_cache))
    logging.info(u'DataSet: {}'.format(list_dataSet))
    logging.info(u'ExcelDatasetFile: {}'.format(list_dataSetexcelfile))
    logging.info(u'Similarity Mode: {}'.format(similarityMode))
    logging.info(u'Embeding File1: {}'.format(os.path.basename(embeddingFile1)))
    logging.info(u'Size of Embeding File1: %.1fMb' % (os.path.getsize(embeddingFile1) / 1024.0 / 1024.0))
    if embeddingFile2 != None:
        logging.info(u'Embeding File2: {}'.format(os.path.basename(embeddingFile2)))
        logging.info(u'Size of Embeding File2: %.1fMb' % (os.path.getsize(embeddingFile2) / 1024.0 / 1024.0))

    global gl_flagPreprocFormsOfWordPair
    global gl_flagtoSelectFromScoreSetofConceptPair
    if similarityMode == SimilarityMode.OnlyUseWordForm:
        logging.info(u'flagPreprocFormsOfWordPair: {}'.format(gl_flagPreprocFormsOfWordPair))
        logging.info(u'flagMultiwordsvecBySumsinglewordvec: {}'.format(gl_flagMultiwordsvecBySumsinglewordvec))
        gl_flagLookupConceptOfWordbyWiki = None #此相似度模式下，不能使用这个参数
        gl_flagtoSelectFromScoreSetofConceptPair = None #此相似度模式下，不能使用这个参数
    elif similarityMode == SimilarityMode.OnlyUseWordConceptFromWikibyConceptVec:
        gl_flagPreprocFormsOfWordPair = None#此相似度模式下，使用的word所对应的wiki concept，不需要做大小写转换
        #logging.info(u'flagPreprocFormsOfWordPair: {} (Forced default value, it is invalid)'.format(gl_flagPreprocFormsOfWordPair))
        assert gl_flagMultiwordsvecBySumsinglewordvec == MultiwordsvecBySumSinglewordvec.DontSumAnywords, \
            u'{}: Now for MultiwordsvecBySumSinglewordvec, I can\'t realize the code for {}'.format(get_cur_info(), gl_flagMultiwordsvecBySumsinglewordvec)
        logging.info(u'flagMultiwordsvecBySumsinglewordvec: {}'.format(gl_flagMultiwordsvecBySumsinglewordvec))
        logging.info(u'flagLookupConceptOfWordbyWiki: {}'.format(gl_flagLookupConceptOfWordbyWiki))
        logging.info(u'flagtoSelectFromScoreSetofConceptPair: {}'.format(gl_flagtoSelectFromScoreSetofConceptPair))
    elif similarityMode == SimilarityMode.UseWordFormAndWordConcept:
        pass
    elif similarityMode == SimilarityMode.OnlyUseWordConceptFromNode2vec:
        gl_flagPreprocFormsOfWordPair = None  # 此相似度模式下，使用的word所对应的wiki concept，不需要做大小写转换
    elif similarityMode == SimilarityMode.OnlyUserCombineWikiConceptNode2vec:
        pass
        # gl_flagPreprocFormsOfWordPair = None  # 此相似度模式下，使用的word所对应的wiki concept，不需要做大小写转换
    else:
        assert False, u'{}: need to update handle code! '.format(get_cur_info())

    logging.info(u'gl_flagtoSelectFromScoreSetofConceptPair: {}'.format(gl_flagtoSelectFromScoreSetofConceptPair))
    logging.info(u'CurrentTime: {}'.format(recTime))
    if advicflagPreprocFormsOfWordPair != gl_flagPreprocFormsOfWordPair:
        logging.info(u'PreprocFormsOfWordPair is set as: {}'.format(gl_flagPreprocFormsOfWordPair))
        logging.info(u'\t whose advice value based on embeddingfile path is: {}'.format(advicflagPreprocFormsOfWordPair))
        logging.info(u'\t they are different, please notice this.')
    time.sleep(0.5)
    echo = raw_input('Please confirm the above parameters! \nInput(y,n):')
    time.sleep(2)
    if echo.lower() != 'y':
        logging.info(u'Please reset the above parameters')
        return

    # 创建缓存文件夹
    if not os.path.exists(gl_cache):
        mkdir_recursively(gl_cache)

    extract_start = default_timer()

    #compute similarity for dataset one by one, Key Code
    list_summarizealldataset = []
    for dataset in list_dataSet:
        logging.info('-----------------------------------')

        ( dataSet, spearman_r, pearson_r, spearman_numofvalidpairs, dataSet_count,
            recTime, recSimilarityMode, recEmbeddingFile1, recEmbeddingFile2, recflagPreprocFormsOfWordPair,
             recflagMultiwordsvecBySumsinglewordvec, recflagLookupConceptOfWordbyWiki, recflagtoSelectFromScoreSetofConceptPair ) \
            = ComputeSimilarityForOneDataSet(dataset, \
                similarityMode,embeddingFile1,embeddingFile2,gl_flagPreprocFormsOfWordPair,\
                gl_flagMultiwordsvecBySumsinglewordvec,recTime, gl_flagLookupConceptOfWordbyWiki, gl_flagtoSelectFromScoreSetofConceptPair)

        list_summarizealldataset.append( ( dataSet, spearman_r, pearson_r, spearman_numofvalidpairs, dataSet_count,
            recTime, recSimilarityMode, recEmbeddingFile1, recEmbeddingFile2, recflagPreprocFormsOfWordPair,
             recflagMultiwordsvecBySumsinglewordvec, recflagLookupConceptOfWordbyWiki, recflagtoSelectFromScoreSetofConceptPair ) )

    #compute a summarized score for assigned dataset and save it to excel
    assert len(list_summarizealldataset)>0
    SaveSummarizedScore2compareexcel(list_summarizealldataset)


    extract_duration = default_timer() - extract_start
    print list_dataSet
    logging.info("\n\n\nFinished computing similarity in %s dataset in %.1f seconds",
                 list_dataSet, extract_duration)

    return

def listitems2strs(sourcelist, splitchars = u'\t'):
    if sourcelist == None:
        return u'None'
    ret = []
    for item in sourcelist:
        ret.append(item)
    ret = splitchars.join(ret)
    return ret


if __name__ == '__main__':
    emb_file = './data/0904005143-CL-CCE_SG_ADD_300_CO0.10_HY0.90.vec'
    pkl_file = './data/enwiki-20161001-pages-articles-multistream-index_IdTitleRedirectNamespaceLemma-Tab-Finalredirect-LowerCompactTitle-pickle.pkl'
    wordpair = Wordpair(emb_file, pkl_file)
    score = wordpair.score('boy', 'lad')
    print score